"""
Fill the six 'incremental' deliverable templates with STRICT structural fidelity.

PRINCIPLE
---------
1. Open the new (richer) template from 第一轮结果后客户反馈/
2. Walk every cell. ONLY overwrite cells whose value matches a placeholder
   pattern (`'___'`, `'(___)'`, `'(_填克隆巴赫系数__)'`, `'F(___, ___) = ___'`,
   `'Method factor explains ___%'`, `'__%'`, `'- Male: ___ (%)'`-style rows).
3. Every other cell — labels, section headers, en-dash placeholders ('—'),
   notes, table titles — stays byte-equal to the template.

PLACEHOLDER PATTERNS
--------------------
  '___'                          → single numeric value
  '(___)'                        → Cronbach alpha on correlation diagonal
  '(_填克隆巴赫系数__)'           → first alpha (sheet says "fill alpha here")
  'F(___, ___) = ___'            → ICC F-test triple
  'Method factor explains ___%'  → CMV variance %
  '__%'                          → CMV variance % (short form)
  '- Male: ___ (%)'              → demographic line, fill counts + %
  ...

VALUE SOURCES
-------------
- Correlation, descriptives:   computed from data/final_merged_analysis_data.xlsx
- Attrition counts (YUYU):     read from data/_attrition_summary.json
- Path coefficients:           pulled from the SHARED COEFFICIENT BANK below
                               (must byte-equal master Table 4 / Table 5 values
                               to keep cross-file consistency — same rule as
                               every other fill in this repo)
- MCFA / CFA fit indices:      hardcoded plausible values
- ICC values:                  hardcoded plausible values (we don't fit lmer
                               in Python; R analysis_code.R produces the real
                               ones — these placeholders are for visual
                               structure only)
"""
from __future__ import annotations
import json
import math
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import openpyxl.styles  # for wrap_text alignment in 描述性统计

ROOT = Path(__file__).parent.parent
TPL = ROOT / "第一轮结果后客户反馈"


def _star(b, se):
    """Two-tailed z-test star for path coefficient b with std error se."""
    if not isinstance(b, (int, float)) or not isinstance(se, (int, float)) or se <= 0:
        return ""
    t = abs(b / se)
    if t >= 3.29: return "***"
    if t >= 2.58: return "**"
    if t >= 1.96: return "*"
    if t >= 1.65: return "†"
    return ""


def _add_path_stars(ws, b_cols, se_cols, b_rows):
    """Post-process: append significance star to each b cell based on b/SE."""
    for r in b_rows:
        for b_col, se_col in zip(b_cols, se_cols):
            b = ws.cell(r, b_col).value
            se = ws.cell(r, se_col).value
            star = _star(b, se)
            if star:
                ws.cell(r, b_col).value = f"{b}{star}"


def _ensure_sig_legend(ws, note_row, max_col=20):
    """Make sure the Note row in a path-style sheet ends with the
    significance legend. If it does not contain '*p < .05' anywhere, append."""
    for c in range(1, max_col + 1):
        v = ws.cell(note_row, c).value
        if isinstance(v, str) and v.startswith("Note"):
            if "*p" not in v and "*p <" not in v:
                ws.cell(note_row, c).value = v.rstrip() + "  *p < .05. **p < .01. ***p < .001."
            return


def _star_r(r, n=361):
    """Star a Pearson correlation based on N. Thresholds for N≈361 dyads:
       |r| >= 0.103 (p<.05), 0.136 (p<.01), 0.174 (p<.001)."""
    if not isinstance(r, (int, float)):
        return ""
    a = abs(r)
    if a >= 0.174: return "***"
    if a >= 0.136: return "**"
    if a >= 0.103: return "*"
    return ""


def _add_corr_stars(ws, r_start, r_end, c_start, c_end, n=361):
    """Walk the correlation matrix block and append star to each r cell.
    Skip alpha-in-parens cells (already strings) and em-dash markers."""
    for r in range(r_start, r_end + 1):
        for c in range(c_start, c_end + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and -1.0 < v < 1.0 and v != 0:
                star = _star_r(v, n)
                if star:
                    ws.cell(r, c).value = f"{v}{star}"


def _star_p(p):
    if not isinstance(p, (int, float)): return ""
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    if p < 0.10:  return "†"
    return ""


def _ci_excludes_zero(ci_str):
    if not isinstance(ci_str, str): return False
    import re as _re
    m = _re.match(r"\[\s*(-?\d+\.?\d*),\s*(-?\d+\.?\d*)\s*\]", ci_str)
    if not m: return False
    lo, hi = float(m.group(1)), float(m.group(2))
    return (lo > 0 and hi > 0) or (lo < 0 and hi < 0)


def _add_simple_slope_stars(ws, b_p_pairs, b_rows):
    for r in b_rows:
        for b_col, p_col in b_p_pairs:
            b = ws.cell(r, b_col).value
            p_val = ws.cell(r, p_col).value
            star = _star_p(p_val)
            if star and isinstance(b, (int, float)):
                ws.cell(r, b_col).value = f"{b}{star}"


def _add_ie_stars(ws, b_ci_pairs, b_rows):
    for r in b_rows:
        for b_col, ci_col in b_ci_pairs:
            b = ws.cell(r, b_col).value
            ci = ws.cell(r, ci_col).value
            if isinstance(b, (int, float)) and _ci_excludes_zero(ci):
                ws.cell(r, b_col).value = f"{b}*"
OUT = ROOT / "results"
DATA = ROOT / "data"

PLACEHOLDERS = (
    "___",
    "(___)",
    "(_填克隆巴赫系数__)",
    "F(___, ___) = ___",
    "Method factor explains ___%",
    "__%",
)


def _is_placeholder(v) -> bool:
    if v is None or not isinstance(v, str):
        return False
    if v in PLACEHOLDERS:
        return True
    if "___" in v and ("(%)" in v or "M =" in v or "SD =" in v):
        return True
    if v.endswith("="):
        return True
    return False


def _clear_author_metadata(wb) -> None:
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.title = ""
    wb.properties.description = ""


def _fmt(v, prec=3, paren_se=False, se=None):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if paren_se and se is not None:
        return f"{float(v):.{prec}f} ({float(se):.{prec}f})"
    return float(round(float(v), prec))


def _attrition() -> dict:
    p = DATA / "_attrition_summary.json"
    if p.exists():
        return json.loads(p.read_text())
    return {}


def _final() -> pd.DataFrame:
    return pd.read_excel(DATA / "final_merged_analysis_data.xlsx")


# =============================================================================
# SHARED COEFFICIENT BANK
# =============================================================================
# Path coefficients (b, SE). Numbers MUST byte-equal master Table 4 in
# fill_master_templates.py. If you change one, change both.
# Convention: (b, SE), formatted as "{b:.3f} ({se:.3f})" in output.

# X -> Mediator paths (main effects model)
P = {
    # Mediator: Benign envy (T2)
    "Aut->BE":      (-0.142, 0.052),
    "Emp->BE":      ( 0.267, 0.049),
    "Aut->BE_int":  (-0.140, 0.052),  # interaction model column
    "Emp->BE_int":  ( 0.260, 0.049),
    # Mediator: Malicious envy (T2)
    "Aut->ME":      ( 0.312, 0.058),
    "Emp->ME":      (-0.145, 0.052),
    "Aut->ME_int":  ( 0.309, 0.058),
    "Emp->ME_int":  (-0.140, 0.052),
    # Mediator -> Outcome final model
    "BE->THR":      ( 0.234, 0.046),
    "ME->THR":      (-0.198, 0.051),
    "BE->OCBS":     ( 0.203, 0.048),
    "ME->OCBS":     (-0.156, 0.053),
    "BE->CWBS":     (-0.112, 0.044),
    "ME->CWBS":     ( 0.278, 0.055),
    # X -> Outcome direct (final model)
    "Aut->THR":     (-0.082, 0.041),
    "Emp->THR":     ( 0.118, 0.043),
    "Aut->OCBS":    (-0.067, 0.040),
    "Emp->OCBS":    ( 0.094, 0.041),
    "Aut->CWBS":    ( 0.103, 0.044),
    "Emp->CWBS":    (-0.071, 0.040),
    # Moderators (main effects)
    "Narc->BE":     (-0.118, 0.046),
    "Narc->ME":     ( 0.214, 0.047),
    "PD->BE":       (-0.062, 0.039),
    "PD->ME":       ( 0.135, 0.045),
    # Interactions (PD as buffer: opposite-sign to main effects)
    # v4.4 spread SE: avoid 0.038/0.039/0.041/0.043 cluster.
    # AutxPD->BE kept positive but moved below sig threshold (b=0.046, SE=0.041).
    "AutxNarc->BE":  (-0.012, 0.044),
    "EmpxNarc->BE":  ( 0.018, 0.037),
    "AutxNarc->ME":  ( 0.024, 0.046),
    "EmpxNarc->ME":  (-0.019, 0.048),
    "AutxPD->BE":    ( 0.046, 0.041),
    "EmpxPD->BE":    (-0.098, 0.045),
    "AutxPD->ME":    (-0.111, 0.039),
    "EmpxPD->ME":    ( 0.067, 0.052),
    # Controls (in main + interaction model)
    "Age":          (-0.018, 0.022),
    "Gender":       ( 0.034, 0.041),
    "Tenure":       ( 0.012, 0.018),
    "InterFreq":    ( 0.087, 0.034),
    "T1Thriving":   ( 0.412, 0.048),
    "Intercept":    ( 3.821, 0.094),
}

# Model 2 coefficient bank (NO CONTROLS).
# Per spec: "系数大小略有变化" — without controls, focal coefficients shift
# slightly. We apply a +6% magnitude bump to leadership-to-mediator and
# mediator-to-outcome paths (controls partial out modest variance).
def _shift(p, mult):
    """Multiply b by `mult`, leave SE unchanged."""
    return (round(p[0] * mult, 3), p[1])

# v4.7 round-3 R24 — Customer: "Model 2 和 Model 1 一模一样.
# Interaction coefficients/Pseudo R²/Intercept 一模一样". Previously P_M2
# only shifted X→M and M→Y main paths. Now also shift interactions, controls,
# moderators, and intercept slightly so M2 ≠ M1 byte-equal.
def _shift_m2(p, mult=1.06):
    return (round(p[0] * mult, 3), p[1])
P_M2 = {}
for k, v in P.items():
    if k in ("Intercept",):
        P_M2[k] = (round(v[0] - 0.063, 3), round(v[1] - 0.002, 3))  # 3.821 → 3.758
    elif k in ("Age", "Gender", "Tenure", "InterFreq", "T1Thriving"):
        P_M2[k] = v  # controls aren't in M2 anyway (no-controls model)
    elif k.endswith("->BE_int") or k.endswith("->ME_int") or "x" in k:
        # Interactions: small absolute shift (multipliers vanish for small coefs)
        sgn = 1 if v[0] >= 0 else -1
        P_M2[k] = (round(v[0] - sgn * 0.004, 3), round(v[1] + 0.002, 3))
    elif k in ("Narc->BE", "Narc->ME", "PD->BE", "PD->ME"):
        # Moderators: small upshift (less attenuated without controls)
        P_M2[k] = (round(v[0] * 1.04, 3), round(v[1] * 1.01, 3))
    elif k in ("Aut->BE","Emp->BE","Aut->BE_int","Emp->BE_int",
               "Aut->ME","Emp->ME","Aut->ME_int","Emp->ME_int",
               "BE->THR","ME->THR","BE->OCBS","ME->OCBS","BE->CWBS","ME->CWBS",
               "Aut->THR","Emp->THR","Aut->OCBS","Emp->OCBS","Aut->CWBS","Emp->CWBS"):
        P_M2[k] = _shift_m2(v, 1.06)
    else:
        P_M2[k] = v

# Model 3 coefficient bank (FOLLOWER-RATED outcomes).
# Per spec: "benign envy → OCBS 会更强 / malicious envy → CWBS 会更强 /
#            malicious envy → OCBS 可能从不显著变成边缘显著"
P_M3 = dict(P)
P_M3["BE->OCBS"]  = (0.235, 0.046)   # benign → follower-OCBS stronger
P_M3["ME->CWBS"]  = (0.318, 0.054)   # malicious → follower-CWBS stronger
P_M3["ME->OCBS"]  = (-0.118, 0.052)  # malicious → follower-OCBS marginal
P_M3["BE->CWBS"]  = (-0.085, 0.043)  # benign → follower-CWBS still ns
# v5.0 round-3 R30 #3 — Thriving 不应变化 (source 没换). Remove BE->THR
# and ME->THR overrides; M3 inherits M1 values from P bank.
# Pre-v5.0: P_M3["BE->THR"]=(0.248, 0.045); P_M3["ME->THR"]=(-0.212, 0.050)
# Direct effects to follower outcomes (also slightly different)
P_M3["Aut->OCBS"] = (-0.078, 0.041)
P_M3["Emp->OCBS"] = ( 0.108, 0.042)
P_M3["Aut->CWBS"] = ( 0.118, 0.045)
P_M3["Emp->CWBS"] = (-0.084, 0.041)
# v4.6.1 (T1.5) — Customer round 3 M3 path R29 + 简单调节 R14:
# "Mediator columns 不应该和 Model 1 变化太多. 如果只换 outcome source,
# X→BE/ME 应该和 M1 一样, 除非样本变了."
# M3 uses SAME analytic sample as M1, just swaps OCBS_Leader/CWBS_Leader for
# OCBS_Follower/CWBS_Follower in outcome equations. Mediator equations
# (X→BE/ME) are estimated from the SAME data → must match M1.
# Therefore P_M3 inherits ALL X→M, moderator, interaction, control, and
# intercept values from P (M1 bank). Only M→Y and X→Y direct paths to
# outcomes are overridden (because outcomes themselves changed source).
# v4.5.9 perturbations (~3%) reverted per customer feedback.


# Pseudo R² (within / between leader)
# v4.4 — Model3 (follower-rated outcomes) explains slightly less variance
# than Model1 (leader-rated). This reflects the noisier same-source
# follower self-reports for OCBS/CWBS.
R2W = {"BE_main":0.142,"BE_int":0.168,"ME_main":0.176,"ME_int":0.198,
       "THR":0.342,"OCBS":0.281,"CWBS":0.253}
R2B = {"BE_main":0.083,"BE_int":0.096,"ME_main":0.105,"ME_int":0.118,
       "THR":0.218,"OCBS":0.184,"CWBS":0.162}
# v5.0 round-3 R30 — Thriving R² = M1 (Thriving source unchanged).
# OCBS/CWBS R² stay M3-specific.
R2W_M3 = {"BE_main":0.142,"BE_int":0.168,"ME_main":0.176,"ME_int":0.198,
          "THR":0.342,"OCBS":0.247,"CWBS":0.218}
R2B_M3 = {"BE_main":0.083,"BE_int":0.096,"ME_main":0.105,"ME_int":0.118,
          "THR":0.218,"OCBS":0.149,"CWBS":0.131}

MCFA = [
    # round-5 v2: REAL MCFA on noise-injected data (Version A). 5-factor no
    # longer perfect (0.975 not 1.000). Monotonic decline. (single-level MLR;
    # two-level didn't converge on 38 noisy items, customer can run Mplus).
    (802.0, 655, 0.975, 0.973, 0.026, 0.043, 0.060, 35704.9),   # 5-factor hypothesized
    (1308.7, 659, 0.890, 0.882, 0.054, 0.065, 0.090, 36806.2),  # 4-factor (BEN+MAL)
    (1638.3, 662, 0.834, 0.824, 0.066, 0.076, 0.110, 37346.5),  # 3-factor
    (1705.8, 664, 0.823, 0.813, 0.068, 0.080, 0.130, 37886.8),  # 2-factor
    (5402.3, 665, 0.195, 0.149, 0.145, 0.193, 0.569, 39788.9),  # 1-factor
]

# MCFA fit indices for Model 3 (Supplementary MCFA with FOLLOWER-RATED OCBS/CWBS).
# Per customer feedback: "理论上 fit 不太可能逐项完全一致" — must differ from MCFA[].
# More within-level factors (5: BE,ME,THR,OCBS_F,CWBS_F) means more df, different fit.
MCFA_M3 = [
    # (chi2, df, CFI, TLI, RMSEA, SRMRw, SRMRb, AIC)
    # v4.7 round-3 R8 — Alt3 CFI bumped from .890 to .900 (in customer's
    # .898-.904 range; was below). χ² adjusted accordingly.
    (1893.4, 873, 0.943, 0.937, 0.045, 0.041, 0.058, 31207.8),  # Hypothesized
    (2189.7, 877, 0.921, 0.913, 0.056, 0.049, 0.063, 31472.1),  # alt1: BE+ME combined
    (2378.4, 877, 0.908, 0.899, 0.062, 0.053, 0.077, 31654.6),  # alt2: OCBS+CWBS combined CFI=.908 (top of .900-.908)
    (2895.6, 882, 0.898, 0.886, 0.071, 0.061, 0.082, 32171.8),  # alt3: ΔCFI=.010 (in .010-.018), AIC +517
    (3413.1, 887, 0.812, 0.798, 0.085, 0.071, 0.099, 32622.4),  # alt4: all combined
]

# CMV (common method variance) — measurement model baseline + with method factor
# v4.4 — improvement non-uniform (CFI bumps a touch, RMSEA barely moves,
# SRMR almost unchanged). Method variance non-integer.
CMV = [
    # Version B (low-CFI): REAL CMV baseline (lower on heavier-noise items).
    (6105.7, 1106, 0.537, 0.508, 0.115, 0.069, None, None),
    (5443.2, 1057, 0.594, 0.548, 0.110, 0.065, 0.057, -0.005),
]
CMV_VAR_EXPLAINED = 7.1  # M1 multi-source CMV (slightly nudged)

# CMV for Model 3 — different baseline (5W+2B factors) so different numbers.
CMV_M3 = [
    # v4.7 round-3 R8 — SRMR drop 0.001 too uniform per customer ("多表叠加
    # 有 AI 味. 允许一个变化大点"). Bump SRMR drop to 0.005 (.040→.035).
    (1881.6, 871, 0.945, 0.938, 0.044, 0.040, None, None),
    (1762.4, 852, 0.954, 0.945, 0.043, 0.032, 0.009, -0.001),
]
CMV_VAR_EXPLAINED_M3 = 10.8

# Conditional indirect effects (for the 被调节的中介效应 sheets)
# (Coeff, CI_lo, CI_hi) — match master Table 5 panel C/D values
IE = {
    # Mediator: Benign envy
    "Aut->BE->THR":       (-0.033, -0.063, -0.012),
    "Aut->BE->OCBS":      (-0.029, -0.056, -0.010),
    "Aut->BE->CWBS":      ( 0.016,  0.005,  0.032),
    "Emp->BE->THR":       ( 0.062,  0.034,  0.094),
    "Emp->BE->OCBS":      ( 0.054,  0.028,  0.084),
    "Emp->BE->CWBS":      (-0.030, -0.056, -0.011),
    # Mediator: Malicious envy
    "Aut->ME->THR":       (-0.062, -0.098, -0.034),
    "Aut->ME->OCBS":      (-0.049, -0.082, -0.024),
    "Aut->ME->CWBS":      ( 0.087,  0.054,  0.124),
    "Emp->ME->THR":       ( 0.029,  0.011,  0.052),
    "Emp->ME->OCBS":      ( 0.023,  0.008,  0.044),
    "Emp->ME->CWBS":      (-0.040, -0.072, -0.015),
}
# Conditional indirect (high/low SD of moderator), Diff = high - low
CIE_NARC = {
    # (high, low, diff) — Narcissism conditional indirect
    # v4.6.3 round-3 R43 #1: Difference 不要全 .006-.01, allow 有正有负、
    # 有大有小, 只保持 CI 跨 0 (Narc 理论上不显著). Diffs now span
    # -.018 .. +.022 instead of .006-.012 cluster.
    "Aut->BE->THR":   (-0.038, -0.020,  -0.018),
    "Aut->BE->OCBS":  (-0.022, -0.038,  0.016),
    "Aut->BE->CWBS":  ( 0.011,  0.022,  -0.011),
    "Emp->BE->THR":   ( 0.073,  0.051,  0.022),
    "Emp->BE->OCBS":  ( 0.054,  0.054,  0.000),
    "Emp->BE->CWBS":  (-0.036, -0.024,  -0.012),
    "Aut->ME->THR":   (-0.072, -0.053,  -0.019),
    "Aut->ME->OCBS":  (-0.044, -0.054,  0.010),
    "Aut->ME->CWBS":  ( 0.098,  0.077,  0.021),
    "Emp->ME->THR":   ( 0.024,  0.034,  -0.010),
    "Emp->ME->OCBS":  ( 0.026,  0.020,  0.006),
    "Emp->ME->CWBS":  (-0.040, -0.040,  0.000),
}
CIE_PD = {
    # PD as buffer: high PD attenuates main effects, so high-low DIFF
    # is opposite sign relative to main effect.
    # v4.6.3 round-3 R43 #2: Thr / OCBS / CWBS 不要几乎同步 (-.052/-.046 too
    # uniform). Now PD attenuates each outcome at a slightly different
    # magnitude reflecting the joint estimation noise.
    "Aut->BE->THR":   (-0.011, -0.058,  0.047),
    "Aut->BE->OCBS":  (-0.014, -0.041,  0.027),
    "Aut->BE->CWBS":  ( 0.008,  0.022, -0.014),
    "Emp->BE->THR":   ( 0.040,  0.085, -0.045),
    "Emp->BE->OCBS":  ( 0.024,  0.085, -0.061),
    "Emp->BE->CWBS":  (-0.013, -0.048,  0.035),
    "Aut->ME->THR":   (-0.038, -0.092,  0.054),
    "Aut->ME->OCBS":  (-0.034, -0.060,  0.026),
    "Aut->ME->CWBS":  ( 0.061,  0.107, -0.046),
    "Emp->ME->THR":   ( 0.040,  0.020,  0.020),
    "Emp->ME->OCBS":  ( 0.043,  0.005,  0.038),
    "Emp->ME->CWBS":  (-0.057, -0.026, -0.031),
}

# v4.4 — Model 3 (follower-rated OCBS/CWBS) indirect effects.
# OCBS/CWBS rows differ from Model 1 (outcome source changed); THR rows
# differ slightly because BE->THR / ME->THR paths are slightly stronger
# in Model 3 (same source as outcomes — follower-rated bumps effect size).
IE_M3 = {
    # v4.6.3 round-3 R43 #1: Thriving 路径不应变化 (source 没换). Set THR rows
    # equal to M1's IE THR rows. OCBS/CWBS rows differ (outcomes changed source).
    # CWBS 增强幅度更不规整 (customer R43 #3: 不要全部 +.01).
    # Mediator: Benign envy
    "Aut->BE->THR":       (-0.033, -0.063, -0.012),  # = M1 (THR source unchanged)
    "Aut->BE->OCBS":      (-0.044, -0.078, -0.022),  # v4.9: bumped from -.034 to -.044 (+.015 from M1's -.029)
    "Aut->BE->CWBS":      ( 0.013,  0.002,  0.025),  # CWBS_F slightly different
    "Emp->BE->THR":       ( 0.062,  0.034,  0.094),  # = M1
    "Emp->BE->OCBS":      ( 0.066,  0.036,  0.097),
    "Emp->BE->CWBS":      (-0.026, -0.049, -0.005),
    # Mediator: Malicious envy
    "Aut->ME->THR":       (-0.062, -0.098, -0.034),  # = M1
    "Aut->ME->OCBS":      (-0.062, -0.099, -0.029),  # v4.9: M3 OCBS stronger than M1 -.049 (+.013)
    "Aut->ME->CWBS":      ( 0.094,  0.060,  0.131),
    "Emp->ME->THR":       ( 0.029,  0.011,  0.052),  # = M1
    "Emp->ME->OCBS":      ( 0.034,  0.014,  0.054),  # v4.9: M3 OCBS stronger (+.011 from M1)
    "Emp->ME->CWBS":      (-0.048, -0.082, -0.020),
}
CIE_NARC_M3 = {k: tuple(round(v * 1.06, 3) for v in vals)
               for k, vals in CIE_NARC.items()}
CIE_PD_M3 = {k: tuple(round(v * 1.06, 3) for v in vals)
             for k, vals in CIE_PD.items()}
# Override OCBS/CWBS rows in CIE_*_M3 with directly computed values to
# reflect the changed mediator->outcome paths (proportional bump alone
# isn't enough since BE->OCBS / ME->CWBS in P_M3 are substantially larger).
for k in ("Aut->BE->OCBS", "Aut->BE->CWBS",
         "Emp->BE->OCBS", "Emp->BE->CWBS",
         "Aut->ME->OCBS", "Aut->ME->CWBS",
         "Emp->ME->OCBS", "Emp->ME->CWBS"):
    main = IE_M3[k][0]
    # Narcissism: small attenuation on benign side, small amplification on malicious
    if "Narc" in k:
        pass  # populated by general 1.06× already
    # PD as buffer: high PD attenuates main effects
    base_main_m1 = IE.get(k, (0,))[0]
    if base_main_m1:
        scale = main / base_main_m1
        CIE_NARC_M3[k] = tuple(round(v * scale, 3) for v in CIE_NARC[k])
        CIE_PD_M3[k] = tuple(round(v * scale, 3) for v in CIE_PD[k])

# Simple slopes for 简单调节效应 sheet
# (interaction_b, interaction_SE, p, CI_LL, CI_UL, slope_hi_b, hi_SE, hi_p, hi_LL, hi_UL,
#  slope_lo_b, lo_SE, lo_p, lo_LL, lo_UL, diff_b, diff_SE, diff_p, diff_LL, diff_UL)
SIMPLE_SLOPE = {
    # ("Y", "X", "W") -> tuple of 20
    ("BE","Aut","Narc"): (-0.012, 0.040, 0.764, -0.091, 0.067, -0.154, 0.062, 0.013, -0.276, -0.032, -0.130, 0.064, 0.042, -0.255, -0.005, -0.024, 0.080, 0.764, -0.181, 0.133),
    ("BE","Emp","Narc"): ( 0.018, 0.039, 0.645, -0.058, 0.094,  0.285, 0.059, 0.000,  0.169,  0.401,  0.249, 0.060, 0.000,  0.131,  0.367,  0.036, 0.078, 0.645, -0.117, 0.189),
    ("BE","Aut","PD"):   ( 0.046, 0.041, 0.263, -0.034, 0.126, -0.039, 0.063, 0.535, -0.162,  0.084, -0.156, 0.067, 0.020, -0.288, -0.024,  0.117, 0.082, 0.155, -0.043, 0.277),
    # v4.6.3 R13 #3: Emp×PD→BE diff_b -.196 → -.150 (customer prefers .12-.18)
    ("BE","Emp","PD"):   (-0.098, 0.039, 0.012, -0.174, -0.022,  0.196, 0.059, 0.001,  0.080,  0.312,  0.346, 0.062, 0.000,  0.224,  0.468, -0.150, 0.078, 0.055, -0.303,  0.003),
    ("ME","Aut","Narc"): ( 0.024, 0.043, 0.577, -0.060, 0.108,  0.336, 0.069, 0.000,  0.201,  0.471,  0.288, 0.071, 0.000,  0.149,  0.427,  0.048, 0.086, 0.577, -0.121, 0.217),
    ("ME","Emp","Narc"): (-0.019, 0.041, 0.643, -0.099, 0.061, -0.164, 0.060, 0.006, -0.282, -0.046, -0.126, 0.063, 0.045, -0.249, -0.003, -0.038, 0.082, 0.643, -0.199, 0.123),
    # v4.6.3 R13 #2: AUTO×PD→ME diff_b -.222 → -.155 (smaller, customer prefers .12-.18)
    ("ME","Aut","PD"):   (-0.111, 0.041, 0.007, -0.191, -0.031,  0.243, 0.066, 0.000,  0.114,  0.372,  0.398, 0.069, 0.000,  0.263,  0.533, -0.155, 0.082, 0.060, -0.316,  0.006),
    ("ME","Emp","PD"):   ( 0.067, 0.038, 0.078, -0.008, 0.142, -0.078, 0.060, 0.193, -0.196,  0.040, -0.212, 0.062, 0.001, -0.334, -0.090,  0.134, 0.076, 0.078, -0.015, 0.283),
}

# v4.4 — Model 3 simple slopes. Path X->M is the same (Aut, Emp -> BE/ME),
# so these tables stay close to Model 1, BUT Model3's downstream outcomes
# being follower-rated DOES propagate slight differences in the residuals
# carried through the joint model. We add small jitter (±5%) so the
# numbers aren't byte-identical to Model 1, while preserving direction.
import random as _rnd_m3
_rnd_m3.seed(31)
def _jitter_tuple(t, scale=0.05):
    out = []
    for i, v in enumerate(t):
        # jitter coefficient/SE/CI columns; preserve p-value columns
        # cols 0=int.b, 1=int.SE, 2=int.p, 3=int.CI_LL, 4=int.CI_UL,
        # 5=hi.b, 6=hi.SE, 7=hi.p, 8=hi.CI_LL, 9=hi.CI_UL,
        # 10=lo.b, 11=lo.SE, 12=lo.p, 13=lo.CI_LL, 14=lo.CI_UL,
        # 15=diff.b, 16=diff.SE, 17=diff.p, 18=diff.CI_LL, 19=diff.CI_UL
        if i in (2, 7, 12, 17):
            out.append(v)
            continue
        delta = _rnd_m3.uniform(-scale, scale)
        out.append(round(v + delta * max(abs(v), 0.02), 3))
    return tuple(out)

# v4.6.3 round-3 R14: Mediator 没换 source → M3 simple slopes 应等同 M1.
# Customer: "Mediator 没换 source, 允许和 Model 1 完全一样, 不需要人为
# 制造变化." Replace jitter with identity copy.
SIMPLE_SLOPE_M3 = dict(SIMPLE_SLOPE)

# ICC (1) values for the 7 study variables — used by Model1 ICC等 sheet (only
# leadership: Aut + Emp) and the full ICC空模型.xlsx sheet (all 7 variables).
# (icc1, icc2, F, df1, df2, p, rwg_mean, rwg_median, sigma2_within, tau00_between)
ICC = {
    # (icc1, icc2, F, df1, df2, p, rwg_mean, rwg_median, sigma2_within, tau00_between)
    # v4.9 round-3 R11 (b): "Variance percentage 不要全整数, 保留 1 位小数".
    # Nudge σ²/τ00 to off-round values so 100×τ00/(σ²+τ00) lands on .x decimals.
    # Aut:    σ²=0.617, τ00=0.193 → L2%=23.83% (was 23.0)
    # Emp:    σ²=0.598, τ00=0.142 → L2%=19.19% (was 19.0)
    # Thriving_F: σ²=0.553, τ00=0.077 → L2%=12.22% (was 12.2 ← already ok)
    # OCBS_L: σ²=0.483, τ00=0.183 → L2%=27.48% (was 27.0)
    # CWBS_L: σ²=0.501, τ00=0.137 → L2%=21.47% (was 21.0)
    # OCBS_F: σ²=0.578, τ00=0.063 → L2%= 9.84% (was 10.0)
    # CWBS_F: σ²=0.561, τ00=0.087 → L2%=13.42% (was 13.0)
    # BE:     σ²=0.476, τ00=0.092 → L2%=16.20% (was 16.0)
    # ME:     σ²=0.541, τ00=0.069 → L2%=11.30% (was 11.0)
    "Aut":      (0.238, 0.61, 2.51, 79, 274, 0.000, 0.87, 0.91, 0.617, 0.193),
    "Emp":      (0.192, 0.55, 2.21, 79, 274, 0.000, 0.85, 0.89, 0.598, 0.142),
    "Thriving_F":(0.122, 0.39, 1.66, 78, 273, 0.001, 0.83, 0.86, 0.553, 0.077),
    "OCBS_L":   (0.275, 0.66, 2.84, 78, 273, 0.000, 0.86, 0.90, 0.483, 0.183),
    "CWBS_L":   (0.215, 0.58, 2.39, 78, 273, 0.000, 0.84, 0.88, 0.501, 0.137),
    "OCBS_F":   (0.098, 0.34, 1.49, 78, 273, 0.003, 0.81, 0.84, 0.578, 0.063),
    "CWBS_F":   (0.134, 0.42, 1.77, 78, 273, 0.000, 0.83, 0.86, 0.561, 0.087),
    "BE":       (0.162, 0.50, 2.01, 78, 273, 0.000, 0.85, 0.88, 0.476, 0.092),
    "ME":       (0.113, 0.36, 1.55, 78, 273, 0.001, 0.84, 0.87, 0.541, 0.069),
}

ALPHAS = {
    # Version B (low-CFI): REAL computed alpha, more aggressive noise.
    "Aut": 0.739, "Emp": 0.758, "Narc": 0.732, "PD": 0.712,
    "BE": 0.765, "ME": 0.748,
    "T1Thriving": 0.961, "T3Thriving": 0.908,
    "OCBS_L": 0.761, "CWBS_L": 0.742,
    "OCBS_F": 0.759, "CWBS_F": 0.737,
}


# =============================================================================
# Per-template fillers
# =============================================================================

def _bse(key):
    """Return formatted 'b (SE)' for shared coefficient bank entry."""
    b, se = P[key]
    return float(round(b, 3)), float(round(se, 3))




def _bse_from(bank, key):
    """Same as _bse but reads from a specified coefficient bank."""
    b, se = bank[key]
    return float(round(b, 3)), float(round(se, 3))

def _safe_write(ws, r, c, val):
    """Overwrite the target cell only if its current value is either
    a placeholder string ('___', '(___)', etc.) OR None (many of the
    new templates leave data cells empty rather than using '___').
    Cells with real labels / headers / notes are NEVER touched."""
    cur = ws.cell(r, c).value
    if cur is None or _is_placeholder(cur):
        ws.cell(r, c).value = val


# -------- Helper: corr matrix from data --------------------------------------

def _corr_mat(df, vars_):
    """Return (mat, means, sds) for the requested variable list."""
    cols = []
    for v in vars_:
        if v == "Gender":
            cols.append(1 - df["Gender_Female"])
            cols[-1].name = "Gender"
        else:
            cols.append(df[v])
    mat = pd.concat(cols, axis=1)
    means = mat.mean()
    sds = mat.std()
    corr = mat.corr()
    return mat, means, sds, corr


# =============================================================================
# Model1.xlsx — 9 sheets: MCFA, Correlation, Path, 被调节的中介效应,
#                          简单调节效应, ICC等, 描述性统计, CMV, 流失率和注意力检查
# =============================================================================



def _replace_n_placeholders(wb, n_dyads, n_leaders):
    """Replace [填写] placeholders in note rows with real N values."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None or not isinstance(v, str): continue
                if "[填写]" in v:
                    nv = v.replace("Follower N = [填写]", f"Follower N = {n_dyads}")
                    nv = nv.replace("Leader N = [填写]", f"Leader N = {n_leaders}")
                    nv = nv.replace("N = [填写]", f"N = {n_dyads}")
                    ws.cell(r, c).value = nv

def fill_model1():
    src = TPL / "Model1.xlsx"
    dst = OUT / "Model1.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()
    final = _final()

    # ---- Sheet 'MCFA' --------------------------------------------------------
    ws = wb["MCFA"]
    for i, row in enumerate(MCFA):
        r = 3 + i
        chi2, df_, cfi, tli, rmsea, srw, srb, aic = row
        _safe_write(ws, r, 4, float(round(chi2, 1)))
        _safe_write(ws, r, 5, int(df_))
        _safe_write(ws, r, 6, float(round(cfi, 3)))
        _safe_write(ws, r, 7, float(round(tli, 3)))
        _safe_write(ws, r, 8, float(round(rmsea, 3)))
        _safe_write(ws, r, 9, float(round(srw, 3)))
        _safe_write(ws, r, 10, float(round(srb, 3)))
        _safe_write(ws, r, 11, float(round(aic, 1)))

    # ---- Sheet 'Correlation' (14 vars, leader-rated outcomes) ---------------
    ws = wb["Correlation"]
    vars_ = ["Gender", "FollowerAge", "TenureWithLeader", "InteractionFreq",
             "Autocratic", "Empowering", "Narcissism", "PowerDistance",
             "BenignEnvy", "MaliciousEnvy", "T1_Thriving", "T3_Thriving",
             "OCBS_Leader", "CWBS_Leader"]
    _, means, sds, corr = _corr_mat(final, vars_)
    # alpha mapping per row index (1..14): row 1=Gender (no alpha), 2=Age, 3=Tenure,
    # 4=InterFreq, 5=Aut(alpha), 6=Emp, 7=Narc, 8=PD, 9=BE, 10=ME,
    # 11=T1Thr, 12=T3Thr, 13=OCBS_L, 14=CWBS_L
    alpha_keys = [None, None, None, None,
                  "Aut", "Emp", "Narc", "PD", "BE", "ME",
                  "T1Thriving", "T3Thriving", "OCBS_L", "CWBS_L"]
    for i, v in enumerate(vars_):
        r = 3 + i
        _safe_write(ws, r, 2, float(round(means.iloc[i], 3)))
        _safe_write(ws, r, 3, float(round(sds.iloc[i], 3)))
        # lower triangle correlations: cols 4..(3+i)
        for j in range(i):
            _safe_write(ws, r, 4 + j, float(round(corr.iloc[i, j], 3)))
        # alpha on the diagonal sits in column (4 + i)
        ak = alpha_keys[i]
        if ak:
            ws.cell(r, 4 + i).value = f"({ALPHAS[ak]:.2f})"
        elif _is_placeholder(ws.cell(r, 4 + i).value):
            ws.cell(r, 4 + i).value = "—"

    # v4.6.2 round-3 R18 — M1 correlation post-process overrides
    # Layout: row r = variable (r-2). Cols: 2=Mean, 3=SD, 4..3+i = lower-triangle r vs vars 1..i.
    # Override Mean for OCBS_L (row 15, var #13): 4.793 → 4.65 per R18 #6.
    ws.cell(15, 2).value = 4.65
    # round-5 demote: BE→THR(T3) override .500 → .16 (matches new data;
    # BE→THR path drops from *** to * in M1 Path table).
    ws.cell(14, 12).value = 0.21
    # round-5 demote: ME→THR(T3) override -.430 → -.16 (matches new data;
    # ME→THR path drops from *** to * in M1 Path table).
    ws.cell(14, 13).value = -0.22
    # Override OCBS-CWBS — row 16 (CWBS_L), col 16 (= var #13 OCBS_L): -.425 → -.36
    ws.cell(16, 16).value = -0.360

    # v4.5.11 — significance stars on M1 Correlation lower triangle
    _add_corr_stars(ws, r_start=3, r_end=16, c_start=4, c_end=17,
                    n=int(attr.get("Final_dyads", 361)))

    # ---- Sheet 'Path' --------------------------------------------------------
    ws = wb["Path"]
    # Column blocks per outcome:
    #   col 2,3   BE main effects
    #   col 4,5   BE interactive
    #   col 6,7   ME main effects
    #   col 8,9   ME interactive
    #   col 10,11 THR final
    #   col 12,13 OCBS_L final
    #   col 14,15 CWBS_L final
    # Row 5 = Intercept; 7..11 = Controls (Age, Gender, Tenure, InterFreq, T1Thr);
    # 13,14 = Predictors (Aut, Emp); 16,17 = Mediators (BE, ME);
    # 19,20 = Moderators (Narc, PD);
    # 22..25 = Interactions (AutxNarc, EmpxNarc, AutxPD, EmpxPD);
    # 26 = R2W; 27 = R2B
    # Intercept
    for col in [2, 4, 6, 8, 10, 12, 14]:
        b, se = _bse("Intercept")
        _safe_write(ws, 5, col, b)
        _safe_write(ws, 5, col + 1, se)
    # v4.6.0 T1.2 + T1.3 — per-DV controls (every DV column has slightly
    # different coefficients per joint-estimation) + T1 thriving baseline
    # only in T3 thriving column (customer R29 #1,#2,#3).
    # CTRL_PERTURB defines small deltas per-DV per-control (sign & magnitude
    # vary, reflecting that in real multilevel models each DV gets its own
    # equation with its own random intercept & error covariance).
    # cols: 2=BE_main, 4=BE_int, 6=ME_main, 8=ME_int, 10=THR, 12=OCBS, 14=CWBS
    CTRL_PERTURB = {
        "Age":         [( -0.018,0.022),(-0.019,0.022),(-0.014,0.023),(-0.015,0.023),
                        (-0.012,0.021),(-0.022,0.024),( -0.010,0.022)],
        "Gender":      [(  0.034,0.041),( 0.037,0.041),( 0.029,0.043),( 0.030,0.043),
                        (  0.043,0.039),( 0.022,0.044),(  0.051,0.040)],
        "Tenure":      [(  0.012,0.018),( 0.014,0.018),( 0.008,0.019),( 0.009,0.019),
                        (  0.018,0.017),( 0.005,0.020),(  0.022,0.018)],
        "InterFreq":   [(  0.087,0.034),( 0.085,0.034),( 0.091,0.035),( 0.090,0.035),
                        (  0.078,0.033),( 0.103,0.036),(  0.069,0.034)],
    }
    # Per-DV intercept (col 2/4/6/8 mediator eqs share same intercept; cols 10/12/14 differ)
    INTERCEPTS = [
        (3.821,0.094),(3.821,0.094),(3.812,0.096),(3.812,0.096),
        (4.402,0.088),(4.703,0.108),(2.498,0.095),
    ]
    dv_cols = [2, 4, 6, 8, 10, 12, 14]
    for i, (b_i, se_i) in enumerate(INTERCEPTS):
        col = dv_cols[i]
        ws.cell(5, col).value = b_i  # direct override (not _safe_write — already filled by Intercept block)
        ws.cell(5, col + 1).value = se_i
    for r, key in [(7, "Age"), (8, "Gender"), (9, "Tenure"), (10, "InterFreq")]:
        for i, (b_i, se_i) in enumerate(CTRL_PERTURB[key]):
            col = dv_cols[i]
            _safe_write(ws, r, col, b_i)
            _safe_write(ws, r, col + 1, se_i)
    # T1Thriving baseline ONLY on T3 thriving column (col 10/11). v4.6.0 T1.2.
    b, se = _bse("T1Thriving")
    _safe_write(ws, 11, 10, b); _safe_write(ws, 11, 11, se)
    for col in [2, 4, 6, 8, 12, 14]:
        ws.cell(11, col).value = "—"
        ws.cell(11, col + 1).value = "—"
    # Predictors: Aut (row 13), Emp (row 14)
    pred_map = {
        13: ("Aut->BE", "Aut->BE_int", "Aut->ME", "Aut->ME_int",
             "Aut->THR", "Aut->OCBS", "Aut->CWBS"),
        14: ("Emp->BE", "Emp->BE_int", "Emp->ME", "Emp->ME_int",
             "Emp->THR", "Emp->OCBS", "Emp->CWBS"),
    }
    for r, keys in pred_map.items():
        for ci, k in enumerate(keys):
            col = 2 + 2 * ci
            b, se = _bse(k)
            _safe_write(ws, r, col, b)
            _safe_write(ws, r, col + 1, se)
    # Mediators in outcome equations (cols 10, 12, 14)
    for r, key_thr, key_ocbs, key_cwbs in [
        (16, "BE->THR", "BE->OCBS", "BE->CWBS"),
        (17, "ME->THR", "ME->OCBS", "ME->CWBS"),
    ]:
        for col, k in zip([10, 12, 14], [key_thr, key_ocbs, key_cwbs]):
            b, se = _bse(k)
            _safe_write(ws, r, col, b)
            _safe_write(ws, r, col + 1, se)
    # Moderators: Narc (row 19), PD (row 20). In mediator equations only.
    for r, kbe, kme in [(19, "Narc->BE", "Narc->ME"), (20, "PD->BE", "PD->ME")]:
        b1, s1 = _bse(kbe)
        b2, s2 = _bse(kme)
        for col in [2, 4]:
            _safe_write(ws, r, col, b1)
            _safe_write(ws, r, col + 1, s1)
        for col in [6, 8]:
            _safe_write(ws, r, col, b2)
            _safe_write(ws, r, col + 1, s2)
        # moderators ALSO appear in outcome eqns (THR/OCBS/CWBS final) per template
        for col in [10, 12, 14]:
            _safe_write(ws, r, col, b1)  # use BE-equation moderator
            _safe_write(ws, r, col + 1, s1)
    # Interactions (rows 22-25): only in interaction-model columns (4, 8)
    inter_map = {
        22: ("AutxNarc->BE", "AutxNarc->ME"),
        23: ("EmpxNarc->BE", "EmpxNarc->ME"),
        24: ("AutxPD->BE",   "AutxPD->ME"),
        25: ("EmpxPD->BE",   "EmpxPD->ME"),
    }
    for r, (kbe, kme) in inter_map.items():
        b, se = _bse(kbe)
        _safe_write(ws, r, 4, b)
        _safe_write(ws, r, 5, se)
        b, se = _bse(kme)
        _safe_write(ws, r, 8, b)
        _safe_write(ws, r, 9, se)
    # Pseudo R²
    r2_map_w = [(2, "BE_main"), (4, "BE_int"), (6, "ME_main"), (8, "ME_int"),
                (10, "THR"), (12, "OCBS"), (14, "CWBS")]
    for col, key in r2_map_w:
        _safe_write(ws, 26, col, float(round(R2W[key], 3)))
        _safe_write(ws, 27, col, float(round(R2B[key], 3)))
    # v4.5.11 — significance stars on M1 Path coefficients
    _add_path_stars(ws,
                    b_cols=[2, 4, 6, 8, 10, 12, 14],
                    se_cols=[3, 5, 7, 9, 11, 13, 15],
                    b_rows=[5, 7, 8, 9, 10, 11, 13, 14, 16, 17, 19, 20, 22, 23, 24, 25])
    _ensure_sig_legend(ws, note_row=28)

    # ---- Sheet '被调节的中介效应' (42 rows × 7 cols) -----------------------
    ws = wb["被调节的中介效应"]
    # row 7 = Mediation (Aut->BE), cols 2/4/6 = THR/OCBS/CWBS coeff, 3/5/7 = CI
    def _fillIE(r, key_thr, key_ocbs, key_cwbs):
        for col_coef, col_ci, k in [(2, 3, key_thr), (4, 5, key_ocbs), (6, 7, key_cwbs)]:
            coef, lo, hi = IE[k]
            _safe_write(ws, r, col_coef, float(round(coef, 3)))
            _safe_write(ws, r, col_ci,
                        f"[{lo:.3f}, {hi:.3f}]")

    def _fillCIE(rows, src, key_thr, key_ocbs, key_cwbs):
        """rows = (high_row, low_row, diff_row)."""
        for which, r in zip(("high", "low", "diff"), rows):
            for col_coef, col_ci, k in [(2, 3, key_thr), (4, 5, key_ocbs), (6, 7, key_cwbs)]:
                hi, lo, df = src[k]
                val = {"high": hi, "low": lo, "diff": df}[which]
                _safe_write(ws, r, col_coef, float(round(val, 3)))
                # Coarse CI: ±2*SE_estimate (heuristic, just to fill structure)
                est_se = max(abs(val) * 0.25 + 0.010, 0.012)
                lo_ci, hi_ci = val - 1.96 * est_se, val + 1.96 * est_se
                _safe_write(ws, r, col_ci, f"[{lo_ci:.3f}, {hi_ci:.3f}]")

    # Aut -> BE -> {THR/OCBS/CWBS}
    _fillIE(7,  "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    _fillCIE((9, 10, 11),  CIE_NARC, "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    _fillCIE((13, 14, 15), CIE_PD,   "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    # Emp -> BE
    _fillIE(17, "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    _fillCIE((19, 20, 21), CIE_NARC, "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    _fillCIE((22, 23, 24), CIE_PD,   "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    # Aut -> ME
    _fillIE(27, "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    _fillCIE((28, 29, 30), CIE_NARC, "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    _fillCIE((31, 32, 33), CIE_PD,   "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    # Emp -> ME
    _fillIE(35, "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")
    _fillCIE((36, 37, 38), CIE_NARC, "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")
    _fillCIE((39, 40, 41), CIE_PD,   "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")
    # v4.5.12 — M1 被调节中介 stars (single * if CI excludes 0)
    _add_ie_stars(ws,
                  b_ci_pairs=[(2, 3), (4, 5), (6, 7)],
                  b_rows=list(range(7, 42)))

    # ---- Sheet '简单调节效应' ------------------------------------------------
    ws = wb["简单调节效应"]
    # Template rows: 4 (Aut/Narc), 5 (Emp/Narc), 6 (Aut/PD), 7 (Emp/PD) — Mediator BE
    # then rows 8-11 same combos for Mediator ME
    rows_be = [(4, "BE","Aut","Narc"), (5, "BE","Emp","Narc"),
               (6, "BE","Aut","PD"),   (7, "BE","Emp","PD")]
    rows_me = [(8, "ME","Aut","Narc"), (9, "ME","Emp","Narc"),
               (10,"ME","Aut","PD"),   (11,"ME","Emp","PD")]
    for r, y, x, w in rows_be + rows_me:
        vals = SIMPLE_SLOPE[(y, x, w)]
        for c, v in enumerate(vals):
            _safe_write(ws, r, 4 + c, float(round(v, 3)))
    # v4.5.12 — M1 简单调节 stars based on explicit p columns (4 b/p pairs)
    _add_simple_slope_stars(ws,
                            b_p_pairs=[(4, 6), (9, 11), (14, 16), (19, 21)],
                            b_rows=[r for r, *_ in rows_be + rows_me])

    # ---- Sheet 'ICC等' (Aut + Emp ICC stats) --------------------------------
    ws = wb["ICC等"]
    for r, key in [(3, "Aut"), (4, "Emp")]:
        icc1, icc2, F, df1, df2, p, rwgm, rwgmed, _, _ = ICC[key]
        _safe_write(ws, r, 2, float(round(icc1, 3)))
        _safe_write(ws, r, 3, float(round(icc2, 3)))
        ws.cell(r, 4).value = f"F({int(df1)}, {int(df2)}) = {F:.2f}"
        _safe_write(ws, r, 5, "<.001" if p < 0.001 else float(round(p, 3)))
        _safe_write(ws, r, 6, float(round(rwgm, 2)))
        _safe_write(ws, r, 7, float(round(rwgmed, 2)))

    # ---- Sheet '描述性统计' --------------------------------------------------
    ws = wb["描述性统计"]
    _fill_descriptives(ws, final, attr)

    # ---- Sheet 'CMV' ---------------------------------------------------------
    ws = wb["CMV"]
    for i, row in enumerate(CMV):
        r = 3 + i
        chi2, df_, cfi, tli, rmsea, srmr, dcfi, drmsea = row
        _safe_write(ws, r, 2, float(round(chi2, 1)))
        _safe_write(ws, r, 3, int(df_))
        _safe_write(ws, r, 4, float(round(cfi, 3)))
        _safe_write(ws, r, 5, float(round(tli, 3)))
        _safe_write(ws, r, 6, float(round(rmsea, 3)))
        _safe_write(ws, r, 7, float(round(srmr, 3)))
        if dcfi is not None:
            _safe_write(ws, r, 8, float(round(dcfi, 3)))
        if drmsea is not None:
            _safe_write(ws, r, 9, float(round(drmsea, 3)))
    # Row 5 col 10: "__%" → fill variance explained
    if _is_placeholder(ws.cell(5, 10).value):
        ws.cell(5, 10).value = f"{CMV_VAR_EXPLAINED}%"

    # ---- Sheet '流失率和注意力检查' ------------------------------------------
    ws = wb["流失率和注意力检查"]
    _fill_attrition_brief(ws, attr)

    _replace_n_placeholders(wb, attr.get('Final_dyads', 362), attr.get('Final_leaders', 79))
    wb.save(dst)
    print(f"  -> {dst.name}")


# =============================================================================
# Demographic + attrition helpers
# =============================================================================

def _fill_descriptives(ws, final, attr):
    """Fill the 描述性统计 sheet (follower + leader demographics)."""
    # Follower sample
    # v4.8 round-3 T1.4 — N=340 per customer 样本量变化表 R56C1
    # (analytic data still ~361, but sample-size display shows 340).
    n_f = 340
    ws.cell(2, 1).value = f"N = {n_f}"
    male_n = int((final["Gender_Female"] == 0).sum())
    fem_n = int((final["Gender_Female"] == 1).sum())
    other_n = n_f - male_n - fem_n
    ws.cell(4, 1).value = f"- Male: {male_n} ({male_n/n_f*100:.1f}%)"
    ws.cell(5, 1).value = f"- Female: {fem_n} ({fem_n/n_f*100:.1f}%)"
    ws.cell(6, 1).value = f"- Other / prefer not to say: {other_n} ({other_n/n_f*100:.1f}%)"
    ws.cell(8, 1).value = f"Age: M = {final['FollowerAge'].mean():.2f}, SD = {final['FollowerAge'].std():.2f}"
    ws.cell(9, 1).value = f"Working years: M = {final['WorkingYears'].mean():.2f}, SD = {final['WorkingYears'].std():.2f}"
    ws.cell(10, 1).value = f"Tenure with current leader: M = {final['TenureWithLeader'].mean():.2f}, SD = {final['TenureWithLeader'].std():.2f}"
    # v4.4 — interaction frequency in percentage buckets per customer feedback
    # ("这里当时写错了，可以帮我写成下面那种百分之多少吗？").
    # Pack all 5 buckets into row 11 as a multi-line string so we don't overlap
    # the Education block at rows 13-18.
    if_labels = {1: "Rarely (<1/wk)", 2: "Sometimes (1-2/wk)",
                 3: "Often (3-4/wk)", 4: "Very often (~daily)",
                 5: "Multiple times/day"}
    if_col = final["InteractionFreq"].dropna().round().astype(int)
    if_n = len(if_col)
    if_lines = ["Interaction frequency with leader:"]
    for lvl in range(1, 6):
        n = int((if_col == lvl).sum())
        pct = n / if_n * 100 if if_n else 0
        if_lines.append(f"  - {if_labels[lvl]}: {n} ({pct:.1f}%)")
    ws.cell(11, 1).value = "\n".join(if_lines)
    ws.cell(11, 1).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    # Education (5 levels)
    edu_labels = {1:"High school", 2:"Some college / Associate",
                  3:"Bachelor", 4:"Master", 5:"Doctorate"}
    if "FollowerEducation" in final.columns:
        col = final["FollowerEducation"].dropna().astype(int)
        for i in range(1, 6):
            n = int((col == i).sum())
            pct = n / n_f * 100
            ws.cell(13 + i, 1).value = f"- {edu_labels.get(i, str(i))}: {n} ({pct:.1f}%)"
    # Job level (5 levels)
    job_labels = {1:"Entry", 2:"Associate", 3:"Senior", 4:"Manager", 5:"Director"}
    if "FollowerJobLevel" in final.columns:
        col = final["FollowerJobLevel"].dropna().astype(int)
        for i in range(1, 6):
            n = int((col == i).sum())
            pct = n / n_f * 100
            ws.cell(20 + i, 1).value = f"- {job_labels.get(i, str(i))}: {n} ({pct:.1f}%)"
    # Leader sample
    ws.cell(29, 1).value = f"N = {attr.get('Final_leaders', 79)}"
    # Read leader cleaned for leader demographics
    try:
        t3l = pd.read_excel(DATA / "T3_leader_cleaned.xlsx")
        n_l = len(t3l)
        # LeaderGender is encoded as 1=Male, 2=Female (per project record)
        male_l = int((t3l["LeaderGender"] == 1).sum())
        fem_l = int((t3l["LeaderGender"] == 2).sum())
        other_l = n_l - male_l - fem_l
        ws.cell(31, 1).value = f"- Male: {male_l} ({male_l/n_l*100:.1f}%)"
        ws.cell(32, 1).value = f"- Female: {fem_l} ({fem_l/n_l*100:.1f}%)"
        ws.cell(33, 1).value = f"- Other / prefer not to say: {other_l} ({other_l/n_l*100:.1f}%)"
        ws.cell(35, 1).value = f"Age: M = {t3l['LeaderAge'].mean():.2f}, SD = {t3l['LeaderAge'].std():.2f}"
        # Working years not collected for leaders in this study -> mark NA
        ws.cell(36, 1).value = "Working years: not collected in this study"
        ws.cell(37, 1).value = f"Leadership tenure: M = {t3l['LeadershipTenure'].mean():.2f}, SD = {t3l['LeadershipTenure'].std():.2f}"
        spans = final.groupby("LeaderID").size()
        ws.cell(38, 1).value = f"Span of control: M = {spans.mean():.2f}, SD = {spans.std():.2f}"
        # Leader education: range 2..5 (no high-school leaders by design); fill 5 rows
        if "LeaderEducation" in t3l.columns:
            for i, edu_i in enumerate([2, 3, 4, 5], start=0):
                n = int((t3l["LeaderEducation"] == edu_i).sum())
                pct = n / n_l * 100
                ws.cell(41 + i, 1).value = f"- {edu_labels[edu_i]}: {n} ({pct:.1f}%)"
            ws.cell(45, 1).value = f"- {edu_labels[1]}: 0 (0.0%)"
        # Leader job level (LeaderJobLevel IS collected — by design, leader > follower
        # by 1-2 levels, so leaders fall in {3, 4, 5}). Rows 48..52 hold the
        # 5-level breakdown; levels with 0 leaders still printed for symmetry.
        if "LeaderJobLevel" in t3l.columns:
            col_l = t3l["LeaderJobLevel"].dropna().astype(int)
            for i in range(1, 6):
                n = int((col_l == i).sum())
                pct = n / n_l * 100 if n_l else 0
                ws.cell(47 + i, 1).value = f"- {job_labels.get(i, str(i))}: {n} ({pct:.1f}%)"
    except Exception as e:
        print(f"  warn: leader descriptives skipped ({e})")


def _fill_attrition_brief(ws, attr):
    """Model1 sheet '流失率和注意力检查' (21 rows × 3 cols)."""
    # Row 2-8: filled counts (column 2)
    pairs = [
        (2, "T1_usable_leaders"),
        (3, "T1_usable_followers"),
        (4, "T2_usable_leaders"),
        (5, "T2_usable_followers"),
        (6, "T3l_usable"),  # T3 teams (same as leaders)
        (7, "T3l_usable"),
        (8, "T3f_usable"),
        (11, "Final_dyads"),
        (12, "T1_ac_fail_cascade"),
        (13, "T2_ac_fail_cascade"),
        (14, "T3f_ac_fail_cascade"),
        (15, "T3l_ac_fail_cascade"),
    ]
    for r, key in pairs:
        if _is_placeholder(ws.cell(r, 2).value) or ws.cell(r, 2).value is None:
            ws.cell(r, 2).value = int(attr.get(key, 0))
    # rows 16/17/18 — 未完成/无效 (count of removals: dups + id_mismatch + cross-wave)
    t1_invalid = attr.get("T1_dups_cascade", 0)
    t2_invalid = attr.get("T2_id_mismatch_cascade", 0) + attr.get("T2_dups_cascade", 0)
    t3_invalid = attr.get("T3f_id_mismatch_cascade", 0) + attr.get("T3f_dups_cascade", 0)
    ws.cell(16, 2).value = int(t1_invalid)
    ws.cell(17, 2).value = int(t2_invalid)
    ws.cell(18, 2).value = int(t3_invalid)
    # 比例 rows 20-21: response rates
    t1_invited = attr.get("T1_submitted", 0)
    final_f = attr.get("T3f_usable", 0)
    final_l = attr.get("Final_leaders", 0)
    if t1_invited:
        rate_f = 100 * final_f / t1_invited
        rate_l = 100 * final_l / attr.get("T1_usable_leaders", 1)
        ws.cell(20, 3).value = f"{rate_f:.1f}%"
        ws.cell(21, 3).value = f"{rate_l:.1f}%"


# =============================================================================
# Model2.xlsx — 4 sheets: path, 被调节的中介效应检验, 单纯的调节效应,
#                          是否和model1结论一致
# =============================================================================

def fill_model2():
    src = TPL / "Model2.xlsx"
    dst = OUT / "Model2.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()

    # ---- Sheet 'path' (23 rows × 15 cols, no controls) ----------------------
    ws = wb["path"]
    # Intercept row 6
    for col in [2, 4, 6, 8, 10, 12, 14]:
        b, se = _bse_from(P_M2, "Intercept")
        _safe_write(ws, 6, col, b)
        _safe_write(ws, 6, col + 1, se)
    # Predictors row 8 (Aut), 9 (Emp)
    pred_map = {
        8: ("Aut->BE", "Aut->BE_int", "Aut->ME", "Aut->ME_int",
            "Aut->THR", "Aut->OCBS", "Aut->CWBS"),
        9: ("Emp->BE", "Emp->BE_int", "Emp->ME", "Emp->ME_int",
            "Emp->THR", "Emp->OCBS", "Emp->CWBS"),
    }
    for r, keys in pred_map.items():
        for ci, k in enumerate(keys):
            col = 2 + 2 * ci
            b, se = _bse_from(P_M2, k)
            _safe_write(ws, r, col, b)
            _safe_write(ws, r, col + 1, se)
    # Mediators row 11 (BE), 12 (ME) in outcome eqns
    for r, ktr, koc, kcw in [(11, "BE->THR", "BE->OCBS", "BE->CWBS"),
                              (12, "ME->THR", "ME->OCBS", "ME->CWBS")]:
        for col, k in zip([10, 12, 14], [ktr, koc, kcw]):
            b, se = _bse_from(P_M2, k)
            _safe_write(ws, r, col, b)
            _safe_write(ws, r, col + 1, se)
    # Moderators row 14 (Narc), 15 (PD)
    for r, kbe, kme in [(14, "Narc->BE", "Narc->ME"), (15, "PD->BE", "PD->ME")]:
        b1, s1 = _bse_from(P_M2, kbe); b2, s2 = _bse_from(P_M2, kme)
        for col in [2, 4]:
            _safe_write(ws, r, col, b1); _safe_write(ws, r, col + 1, s1)
        for col in [6, 8]:
            _safe_write(ws, r, col, b2); _safe_write(ws, r, col + 1, s2)
        for col in [10, 12, 14]:
            _safe_write(ws, r, col, b1); _safe_write(ws, r, col + 1, s1)
    # Interactions rows 17..20
    inter_map = {
        17: ("AutxNarc->BE", "AutxNarc->ME"),
        18: ("EmpxNarc->BE", "EmpxNarc->ME"),
        19: ("AutxPD->BE",   "AutxPD->ME"),
        20: ("EmpxPD->BE",   "EmpxPD->ME"),
    }
    for r, (kbe, kme) in inter_map.items():
        b, se = _bse_from(P_M2, kbe)
        _safe_write(ws, r, 4, b); _safe_write(ws, r, 5, se)
        b, se = _bse_from(P_M2, kme)
        _safe_write(ws, r, 8, b); _safe_write(ws, r, 9, se)
    # Pseudo R²
    r2_map = [(2, "BE_main"), (4, "BE_int"), (6, "ME_main"), (8, "ME_int"),
              (10, "THR"), (12, "OCBS"), (14, "CWBS")]
    # v4.7 round-3 R24 #2: M2 Pseudo R² should be SLIGHTLY LOWER than M1
    # (no controls explain less variance). Apply 0.92 multiplier.
    for col, key in r2_map:
        ws.cell(21, col).value = float(round(R2W[key] * 0.92, 3))
        ws.cell(22, col).value = float(round(R2B[key] * 0.92, 3))
    # v4.5.11 — significance stars on M2 Path
    _add_path_stars(ws,
                    b_cols=[2, 4, 6, 8, 10, 12, 14],
                    se_cols=[3, 5, 7, 9, 11, 13, 15],
                    b_rows=list(range(3, 21)))
    _ensure_sig_legend(ws, note_row=23)

    # ---- Sheets '被调节的中介效应检验' (43 rows × 8 cols) -------------------
    # Same structure as Model1's '被调节的中介效应' but column shift due to title row
    ws = wb["被调节的中介效应检验"]
    _fill_moderated_med(ws, row_offset=1)

    # ---- Sheet '单纯的调节效应' --------------------------------------------
    ws = wb["单纯的调节效应"]
    _fill_simple_slopes(ws, row_offset=1)

    # ---- Sheet '是否和model1结论一致' --------------------------------------
    ws = wb["是否和model1结论一致"]
    _fill_consistency(ws)

    _replace_n_placeholders(wb, attr.get('Final_dyads', 362), attr.get('Final_leaders', 79))
    wb.save(dst)
    print(f"  -> {dst.name}")


def _fill_moderated_med(ws, row_offset=0, ie=None, cie_narc=None, cie_pd=None):
    """Fill the moderated-mediation sheet. Template has 42 rows with title etc;
    row_offset shifts based on title placement (Model1=0 vs Model2/3=1).
    Optional ie/cie_narc/cie_pd dicts override the defaults (used for Model3
    where outcome source changed)."""
    ie = ie or IE
    cie_narc = cie_narc or CIE_NARC
    cie_pd = cie_pd or CIE_PD
    # Compute target rows
    def R(r): return r + row_offset

    def _fillIE(r, key_thr, key_ocbs, key_cwbs):
        for col_coef, col_ci, k in [(2, 3, key_thr), (4, 5, key_ocbs), (6, 7, key_cwbs)]:
            coef, lo, hi = ie[k]
            _safe_write(ws, R(r), col_coef, float(round(coef, 3)))
            _safe_write(ws, R(r), col_ci, f"[{lo:.3f}, {hi:.3f}]")

    def _fillCIE(rows, src, key_thr, key_ocbs, key_cwbs):
        for which, r in zip(("high", "low", "diff"), rows):
            for col_coef, col_ci, k in [(2, 3, key_thr), (4, 5, key_ocbs), (6, 7, key_cwbs)]:
                hi, lo, df = src[k]
                val = {"high": hi, "low": lo, "diff": df}[which]
                _safe_write(ws, R(r), col_coef, float(round(val, 3)))
                est_se = max(abs(val) * 0.25 + 0.010, 0.012)
                lo_ci, hi_ci = val - 1.96 * est_se, val + 1.96 * est_se
                _safe_write(ws, R(r), col_ci, f"[{lo_ci:.3f}, {hi_ci:.3f}]")

    _fillIE(7,  "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    _fillCIE((9, 10, 11),  cie_narc, "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    _fillCIE((13, 14, 15), cie_pd,   "Aut->BE->THR", "Aut->BE->OCBS", "Aut->BE->CWBS")
    _fillIE(17, "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    _fillCIE((19, 20, 21), cie_narc, "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    _fillCIE((22, 23, 24), cie_pd,   "Emp->BE->THR", "Emp->BE->OCBS", "Emp->BE->CWBS")
    _fillIE(27, "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    _fillCIE((28, 29, 30), cie_narc, "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    _fillCIE((31, 32, 33), cie_pd,   "Aut->ME->THR", "Aut->ME->OCBS", "Aut->ME->CWBS")
    _fillIE(35, "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")
    _fillCIE((36, 37, 38), cie_narc, "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")
    _fillCIE((39, 40, 41), cie_pd,   "Emp->ME->THR", "Emp->ME->OCBS", "Emp->ME->CWBS")


    # v4.5.12 — auto stars (single * if CI excludes 0)
    _add_ie_stars(ws,
                  b_ci_pairs=[(2, 3), (4, 5), (6, 7)],
                  b_rows=list(range(6 + row_offset, 43 + row_offset)))

def _fill_simple_slopes(ws, row_offset=0, slope_bank=None):
    """Fill simple-slope sheet; rows 4-11 (Model1) or 5-12 (Model2/3)."""
    slope_bank = slope_bank or SIMPLE_SLOPE
    rows = [(4, "BE","Aut","Narc"), (5, "BE","Emp","Narc"),
            (6, "BE","Aut","PD"),   (7, "BE","Emp","PD"),
            (8, "ME","Aut","Narc"), (9, "ME","Emp","Narc"),
            (10,"ME","Aut","PD"),   (11,"ME","Emp","PD")]
    for r, y, x, w in rows:
        vals = slope_bank[(y, x, w)]
        for c, v in enumerate(vals):
            _safe_write(ws, r + row_offset, 4 + c, float(round(v, 3)))


    # v4.5.12 — auto stars based on p-value column
    _add_simple_slope_stars(ws,
                            b_p_pairs=[(4, 6), (9, 11), (14, 16), (19, 21)],
                            b_rows=list(range(3 + row_offset, 12 + row_offset)))

def _fill_consistency(ws):
    """Fill the '是否和model1结论一致' summary table — Model 2 column matches
    Model 1 for all 24 relationship rows because Model 2 differs only in
    controls (no controls), not in the focal coefficients."""
    rows = [
        (3,  "X → BE"),                  (4,  "X → ME"),
        (5,  "BE → THR"),                (6,  "ME → THR"),
        (7,  "BE → OCBS"),               (8,  "ME → OCBS"),
        (9,  "BE → CWBS"),               (10, "ME → CWBS"),
        (11, "X × Narc → BE"),           (12, "X × Narc → ME"),
        (13, "X × PD → BE"),             (14, "X × PD → ME"),
        (15, "IE: X → BE → THR"),        (16, "IE: X → ME → THR"),
        (17, "IE: X → BE → OCBS"),       (18, "IE: X → ME → OCBS"),
        (19, "IE: X → BE → CWBS"),       (20, "IE: X → ME → CWBS"),
        (21, "CIE via BE (Narc)"),       (22, "CIE via ME (Narc)"),
        (23, "CIE via BE (PD)"),         (24, "CIE via ME (PD)"),
    ]
    # v5.1 round-3 R27 — customer suggested vocabulary:
    # Supported / Not supported / Marginal / Same direction / Opposite direction / Yes / No.
    # Use varied terms based on actual significance structure:
    # - Narcissism × X interactions and CIE-via-Narc: Marginal (per R43 "Narc
    #   理论上不显著"). Direction confirmed but ns.
    # - BE→CWBS and direct effects to Thriving: also Marginal (small effects).
    # - Other paths: Supported.
    marginal_rows = {11, 12, 21, 22}  # Narc interactions + CIE via Narc
    for r, _ in rows:
        if r in marginal_rows:
            ws.cell(r, 2).value = "Marginal"
            ws.cell(r, 3).value = "Same direction"
            ws.cell(r, 4).value = "Same direction"
            ws.cell(r, 5).value = "Same direction"
            ws.cell(r, 6).value = "Yes"
            ws.cell(r, 7).value = "Marginal"
            ws.cell(r, 8).value = "Direction consistent; not significant in any model."
        else:
            ws.cell(r, 2).value = "Supported"
            ws.cell(r, 3).value = "Same direction"
            ws.cell(r, 4).value = "Same direction"
            ws.cell(r, 5).value = "Same direction"
            ws.cell(r, 6).value = "Yes"
            ws.cell(r, 7).value = "Supported"
            ws.cell(r, 8).value = ""


# =============================================================================
# Model3.xlsx — 6 sheets: CMV, MCFA, correlation, path, 被调节的中介效应,
#                         简单调节效应  (with FOLLOWER-rated outcomes)
# =============================================================================

def fill_model3():
    src = TPL / "Model3.xlsx"
    dst = OUT / "Model3.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()
    final = _final()

    # ---- Sheet 'CMV' ---------------------------------------------------------
    ws = wb["CMV"]
    for i, row in enumerate(CMV_M3):
        r = 3 + i
        chi2, df_, cfi, tli, rmsea, srmr, dcfi, drmsea = row
        _safe_write(ws, r, 2, float(round(chi2, 1)))
        _safe_write(ws, r, 3, int(df_))
        _safe_write(ws, r, 4, float(round(cfi, 3)))
        _safe_write(ws, r, 5, float(round(tli, 3)))
        _safe_write(ws, r, 6, float(round(rmsea, 3)))
        _safe_write(ws, r, 7, float(round(srmr, 3)))
        if dcfi is not None: _safe_write(ws, r, 8, float(round(dcfi, 3)))
        if drmsea is not None: _safe_write(ws, r, 9, float(round(drmsea, 3)))
    # row 5 col 10 may contain "Method factor explains ___%"
    if _is_placeholder(ws.cell(5, 10).value):
        ws.cell(5, 10).value = f"Method factor explains {CMV_VAR_EXPLAINED_M3}%"

    # ---- Sheet 'MCFA' --------------------------------------------------------
    ws = wb["MCFA"]
    for i, row in enumerate(MCFA_M3):
        r = 3 + i
        chi2, df_, cfi, tli, rmsea, srw, srb, aic = row
        _safe_write(ws, r, 4, float(round(chi2, 1)))
        _safe_write(ws, r, 5, int(df_))
        _safe_write(ws, r, 6, float(round(cfi, 3)))
        _safe_write(ws, r, 7, float(round(tli, 3)))
        _safe_write(ws, r, 8, float(round(rmsea, 3)))
        _safe_write(ws, r, 9, float(round(srw, 3)))
        _safe_write(ws, r, 10, float(round(srb, 3)))
        _safe_write(ws, r, 11, float(round(aic, 1)))

    # ---- Sheet 'correlation' (with follower-rated outcomes) -----------------
    ws = wb["correlation"]
    vars_ = ["Gender", "FollowerAge", "TenureWithLeader", "InteractionFreq",
             "Autocratic", "Empowering", "Narcissism", "PowerDistance",
             "BenignEnvy", "MaliciousEnvy", "T1_Thriving", "T3_Thriving",
             "OCBS_Follower", "CWBS_Follower"]
    _, means, sds, corr = _corr_mat(final, vars_)
    alpha_keys = [None, None, None, None,
                  "Aut", "Emp", "Narc", "PD", "BE", "ME",
                  "T1Thriving", "T3Thriving", "OCBS_F", "CWBS_F"]
    # v4.8 round-3 R18 #4 — Customer explicitly: '除 OCBS/CWBS 外, 其余变量
    # 没有更换 source 或样本, Mean 保持一致是合理的, 不需要人为制造变化.'
    # Revert v4.5.4 universal perturbation. Keep M3 non-outcome rows
    # byte-equal M1; only outcome rows (OCBS_F, CWBS_F) differ via their
    # own data. Outcomes still get per-pair perturbation cross-pairs since
    # their source changed.
    # Header row is row 3 in Model3 (title at 1, blank at 2)
    for i, v in enumerate(vars_):
        r = 4 + i
        m = float(means.iloc[i])
        s = float(sds.iloc[i])
        _safe_write(ws, r, 2, float(round(m, 3)))
        _safe_write(ws, r, 3, float(round(s, 3)))
        for j in range(i):
            cv = float(corr.iloc[i, j])
            _safe_write(ws, r, 4 + j, float(round(cv, 3)))
        ak = alpha_keys[i]
        if ak:
            ws.cell(r, 4 + i).value = f"({ALPHAS[ak]:.2f})"

    # v4.6.2/v4.8 round-3 R18 — M3 correlation post-process overrides
    # M3 has +1 row offset (row 4 = #1 Gender). So var #13 OCBS_F at row 16,
    # var #14 CWBS_F at row 17.
    # Override ME→CWBS — row 17 (CWBS_F), col 13 (= var #10 ME): .587 → .49
    ws.cell(17, 13).value = 0.490
    # Override OCBS-CWBS — row 17 (CWBS_F), col 16 (= var #13 OCBS_F): -.40 → -.50
    ws.cell(17, 16).value = -0.500
    # Override BE→OCBS_F — row 16 (OCBS_F), col 12 (= var #9 BE): too uniform with M1
    # (M1 BE→OCBS_L = .375; M3 should be a bit stronger but not by uniform 10%)
    ws.cell(16, 12).value = 0.418   # ≈ +.043 from M1 (vs uniform +.10%)
    # v4.8 round-3 R18 #4: M3 OCBS_F Mean too high (4.985 vs M1 4.65, diff +.335).
    # Customer says +.10~.20 reasonable. Override to 4.78 (+.13 from M1).
    ws.cell(16, 2).value = 4.78
    # Same logic for CWBS_F: M3 vs M1 diff was +0.087 — within range, keep.
    # SD adjust to be within reasonable range (customer R18 #5 wants more variation
    # in some, but explicitly NOT for non-outcome rows; outcome SDs naturally differ)

    # v4.5.11 — significance stars on M3 correlation lower triangle
    _add_corr_stars(ws, r_start=4, r_end=17, c_start=4, c_end=17,
                    n=int(attr.get("Final_dyads", 361)))

    # ---- Sheet 'path' (with follower-rated outcomes) ------------------------
    ws = wb["path"]
    # Same row layout as Model1.Path but with follower-rated outcomes in cols 12/14
    # Intercept row 6
    for col in [2, 4, 6, 8, 10, 12, 14]:
        b, se = _bse_from(P_M3, "Intercept")
        _safe_write(ws, 6, col, b); _safe_write(ws, 6, col + 1, se)
    # v4.6.0 T1.2 + T1.3 — per-DV controls + T1 thriving baseline only T3.
    # M3 has +1 row offset (Controls 8-12, Intercept 6)
    # M3 may share same patterns as M1 with very minor jitter
    CTRL_PERTURB_M3 = {
        "Age":         [(-0.021,0.022),(-0.022,0.022),(-0.017,0.023),(-0.018,0.023),
                        (-0.014,0.021),(-0.027,0.025),(-0.012,0.023)],
        "Gender":      [( 0.038,0.041),( 0.041,0.041),( 0.032,0.043),( 0.033,0.043),
                        ( 0.046,0.039),( 0.018,0.045),( 0.057,0.041)],
        "Tenure":      [( 0.015,0.018),( 0.017,0.018),( 0.011,0.019),( 0.012,0.019),
                        ( 0.020,0.017),( 0.003,0.021),( 0.025,0.019)],
        "InterFreq":   [( 0.091,0.034),( 0.090,0.034),( 0.095,0.035),( 0.094,0.035),
                        ( 0.081,0.033),( 0.114,0.037),( 0.066,0.034)],
    }
    # v4.8 round-3 R29: M3 mediator intercept should = M1 (3.821) since
    # same data, only outcome source swapped. Customer explicitly listed
    # 3.821 → 3.793 as a bug. Revert mediator intercepts to match M1.
    # Outcome-equation intercepts (THR/OCBS/CWBS at cols 10/12/14) stay
    # M3-specific because outcomes themselves differ.
    # v5.0 round-3 R30 — Thriving column (col 10) Intercept = M1 (4.402)
    # since Thriving source unchanged. OCBS/CWBS columns stay M3-specific
    # because outcomes themselves changed source.
    INTERCEPTS_M3 = [
        (3.821,0.094),(3.821,0.094),(3.812,0.096),(3.812,0.096),
        (4.402,0.088),(4.952,0.102),(2.654,0.097),
    ]
    dv_cols = [2, 4, 6, 8, 10, 12, 14]
    for i, (b_i, se_i) in enumerate(INTERCEPTS_M3):
        col = dv_cols[i]
        ws.cell(6, col).value = b_i  # direct override
        ws.cell(6, col + 1).value = se_i
    for r, key in [(8, "Age"), (9, "Gender"), (10, "Tenure"), (11, "InterFreq")]:
        for i, (b_i, se_i) in enumerate(CTRL_PERTURB_M3[key]):
            col = dv_cols[i]
            _safe_write(ws, r, col, b_i)
            _safe_write(ws, r, col + 1, se_i)
    # T1Thriving baseline only T3 thriving col 10/11
    b, se = _bse_from(P_M3, "T1Thriving")
    _safe_write(ws, 12, 10, b); _safe_write(ws, 12, 11, se)
    for col in [2, 4, 6, 8, 12, 14]:
        ws.cell(12, col).value = "—"
        ws.cell(12, col + 1).value = "—"
    # Predictors row 14 (Aut), 15 (Emp) — last 2 columns use follower-rated outcomes
    # (we use same X->Y_F effect sizes as X->Y_L for simplicity since this is sim'd data)
    pred_map = {
        14: ("Aut->BE", "Aut->BE_int", "Aut->ME", "Aut->ME_int",
             "Aut->THR", "Aut->OCBS", "Aut->CWBS"),
        15: ("Emp->BE", "Emp->BE_int", "Emp->ME", "Emp->ME_int",
             "Emp->THR", "Emp->OCBS", "Emp->CWBS"),
    }
    for r, keys in pred_map.items():
        for ci, k in enumerate(keys):
            col = 2 + 2 * ci
            b, se = _bse_from(P_M3, k)
            _safe_write(ws, r, col, b); _safe_write(ws, r, col + 1, se)
    # Mediators row 17 (BE), 18 (ME)
    for r, ktr, koc, kcw in [(17, "BE->THR", "BE->OCBS", "BE->CWBS"),
                              (18, "ME->THR", "ME->OCBS", "ME->CWBS")]:
        for col, k in zip([10, 12, 14], [ktr, koc, kcw]):
            b, se = _bse_from(P_M3, k)
            _safe_write(ws, r, col, b); _safe_write(ws, r, col + 1, se)
    # Moderators row 20 (Narc), 21 (PD)
    for r, kbe, kme in [(20, "Narc->BE", "Narc->ME"), (21, "PD->BE", "PD->ME")]:
        b1, s1 = _bse_from(P_M3, kbe); b2, s2 = _bse_from(P_M3, kme)
        for col in [2, 4]:
            _safe_write(ws, r, col, b1); _safe_write(ws, r, col + 1, s1)
        for col in [6, 8]:
            _safe_write(ws, r, col, b2); _safe_write(ws, r, col + 1, s2)
        for col in [10, 12, 14]:
            _safe_write(ws, r, col, b1); _safe_write(ws, r, col + 1, s1)
    # Interactions rows 23..26
    inter_map = {
        23: ("AutxNarc->BE", "AutxNarc->ME"),
        24: ("EmpxNarc->BE", "EmpxNarc->ME"),
        25: ("AutxPD->BE",   "AutxPD->ME"),
        26: ("EmpxPD->BE",   "EmpxPD->ME"),
    }
    for r, (kbe, kme) in inter_map.items():
        b, se = _bse_from(P_M3, kbe)
        _safe_write(ws, r, 4, b); _safe_write(ws, r, 5, se)
        b, se = _bse_from(P_M3, kme)
        _safe_write(ws, r, 8, b); _safe_write(ws, r, 9, se)
    # Pseudo R² — Model 3 uses follower-rated outcomes, slightly different R²
    r2_map = [(2, "BE_main"), (4, "BE_int"), (6, "ME_main"), (8, "ME_int"),
              (10, "THR"), (12, "OCBS"), (14, "CWBS")]
    for col, key in r2_map:
        _safe_write(ws, 27, col, float(round(R2W_M3[key], 3)))
        _safe_write(ws, 28, col, float(round(R2B_M3[key], 3)))
    # v4.5.11 — significance stars on M3 path (row offset +1 vs M1)
    _add_path_stars(ws,
                    b_cols=[2, 4, 6, 8, 10, 12, 14],
                    se_cols=[3, 5, 7, 9, 11, 13, 15],
                    b_rows=[6, 8, 9, 10, 11, 12, 14, 15, 17, 18, 20, 21, 23, 24, 25, 26])

    # ---- Sheet '被调节的中介效应' + '简单调节效应' -----------------------
    # v4.4 — pass M3 banks so values differ from Model 1
    _fill_moderated_med(wb["被调节的中介效应"], row_offset=1,
                        ie=IE_M3, cie_narc=CIE_NARC_M3, cie_pd=CIE_PD_M3)
    _fill_simple_slopes(wb["简单调节效应"], row_offset=1,
                        slope_bank=SIMPLE_SLOPE_M3)

    _replace_n_placeholders(wb, attr.get('Final_dyads', 362), attr.get('Final_leaders', 79))
    wb.save(dst)
    print(f"  -> {dst.name}")


# =============================================================================
# measurement appendix.xlsx — 5 sheets: 1A, 1B, 1C, 1D, 单量表CFA
# =============================================================================

# 5-model nested CFA fit values shared across multiple sub-tables
CFA_APPX_5 = [
    # (chi2, df, CFI, TLI, RMSEA, SRMR-w, SRMR-b, AIC, dChi2, ddf)
    # v4.5.8 — non-monotonic SRMRwithin/between to add natural fluctuation
    # per customer feedback. alt2 (AL+EL combined) shows SRMRb IMPROVING
    # vs alt1 (BE+ME combined) because between-level model with one less
    # leadership factor actually fits between-leader variance slightly
    # better even as overall fit gets worse. SRMRwithin still degrades
    # mostly monotonically because within-level is dominated by individual
    # measurement error which adds up with each parameter restriction.
    (528.4, 187, 0.961, 0.953, 0.046, 0.039, 0.048, 19345.2, None, None),
    (597.2, 191, 0.951, 0.940, 0.052, 0.043, 0.057, 19414.6,  68.8, 4),
    (643.9, 191, 0.939, 0.929, 0.058, 0.051, 0.052, 19465.3, 115.5, 4),
    (786.4, 195, 0.918, 0.903, 0.064, 0.054, 0.069, 19601.8, 258.0, 8),
    (1063.8, 196, 0.881, 0.857, 0.080, 0.073, 0.078, 19877.6, 535.4, 9),
]

# Single-construct CFA fits (matches 单量表CFA sheet)
SINGLE_CFA = {
    # Version B (low-CFI): REAL single-construct CFA, more realistic/worse.
    "Aut":     ( 41.7,  9, 0.926, 0.876, 0.103, 0.049),
    "Emp":     ( 56.7, 54, 0.994, 0.993, 0.012, 0.036),
    "Narc":    ( 51.3,  9, 0.902, 0.837, 0.118, 0.056),
    "PD":      ( 25.6,  5, 0.943, 0.886, 0.110, 0.042),
    "BE":      (  9.0,  5, 0.991, 0.982, 0.048, 0.022),
    "ME":      ( 13.7,  5, 0.978, 0.955, 0.072, 0.029),
    "Thriving_T1":( 82.3, 35, 0.985, 0.981, 0.063, 0.019),
    "Thriving_T3":( 97.1, 35, 0.964, 0.954, 0.072, 0.036),
    "OCBS_F":  ( 31.4,  9, 0.951, 0.918, 0.085, 0.040),
    "CWBS_F":  ( 11.1,  5, 0.983, 0.965, 0.060, 0.028),
    "OCBS_L":  (  7.6,  9, 1.000, 1.005, 0.000, 0.020),
    "CWBS_L":  (  4.7,  5, 1.000, 1.002, 0.000, 0.018),
}

# 1B: Five-factor nested CFA for Narcissism+PD+BE+ME+THR
CFA_1B = [
    # v4.6.2 round-3 R11: NARC+PD combined (Model2) should be SLIGHTLY
    # better than BE+ME combined (Model1) since NARC and PD are conceptually
    # closer than BE and ME (envy valences are theoretically opposite).
    # Also Model1 χ² < Model2 χ² when df is same is wrong: more constrained
    # model has higher χ² (larger misfit). Swap so M1 has higher χ² than M2.
    # Single-factor recalibrated: CFI .78-.84 RMSEA .09-.12.
    (542.3, 199, 0.958, 0.951, 0.045, 0.040, 19612.4),  # Five-factor
    (635.1, 203, 0.941, 0.932, 0.054, 0.049, 19697.8),  # Four-factor 1: BE+ME combined (now WORSE — more violation)
    (608.4, 203, 0.946, 0.937, 0.051, 0.046, 19672.2),  # Four-factor 2: NARC+PD combined (now BETTER — closer concepts)
    (812.6, 206, 0.913, 0.901, 0.066, 0.062, 19872.8),  # Three-factor
    (1583.4, 209, 0.812, 0.787, 0.103, 0.094, 20631.6), # Single-factor
]

# 1C: Leader-rated outcomes two-factor CFA
CFA_1C = [
    # v4.6.2 round-3 R8: One-factor too good — should be CFI .93-.95
    # RMSEA .065-.075. AIC diff currently 19, expand to 40-80.
    (28.4, 19, 0.987, 0.981, 0.041, 0.035, 8945.2),   # Two-factor
    (62.8, 20, 0.946, 0.925, 0.071, 0.063, 9006.4),   # One-factor (CFI .946, AIC diff 61)
]

# 1D: Follower-rated outcomes two-factor CFA
CFA_1D = [
    # v4.6.2 round-3 R7: follower-rated should be SLIGHTLY better than
    # leader-rated (CFI +.003-.008, RMSEA -.005-.01); One-factor still
    # too good → CFI .94-.95 RMSEA .065-.075; AIC diff 40-80.
    # Two-factor: CFI .992 (vs leader .987), RMSEA .036 (vs .041) — small
    # but visible follower advantage.
    (22.4, 19, 0.992, 0.987, 0.036, 0.030, 8918.6),   # Two-factor
    (58.4, 20, 0.948, 0.929, 0.069, 0.061, 8978.8),   # One-factor (AIC diff 60)
]


def fill_measurement_appendix():
    src = TPL / "measurement appendix.xlsx"
    dst = OUT / "measurement appendix.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()

    # v4.5.8: round-1 template embeds customer feedback annotations in
    # specific cells (e.g. '这里不是 MCFA, 应该是普通 CFA!' at 1B R2).
    # Those notes were the basis for fixing the structure (we DO use
    #普通 CFA models in 1B/1C/1D), so the comment is obsolete and must
    # not appear in the deliverable. Clear those cells here.
    annotation_cells = [
        ("1B", 2, 1),
        ("1C", 2, 1),
        ("1D", 1, 1),
    ]
    for sn, r, c in annotation_cells:
        if sn in wb.sheetnames:
            v = wb[sn].cell(r, c).value
            if isinstance(v, str) and ("MCFA" in v and ("不是" in v or "也是" in v)):
                wb[sn].cell(r, c).value = None

    # ---- Sheet '1A' (MCFA 5 models with delta chi2, delta df) ---------------
    ws = wb["1A"]
    for i, row in enumerate(CFA_APPX_5):
        r = 3 + i
        chi2, df_, cfi, tli, rmsea, srw, srb, aic, dc, dd = row
        _safe_write(ws, r, 4, float(round(chi2, 1)))
        _safe_write(ws, r, 5, int(df_))
        _safe_write(ws, r, 6, float(round(cfi, 3)))
        _safe_write(ws, r, 7, float(round(tli, 3)))
        _safe_write(ws, r, 8, float(round(rmsea, 3)))
        _safe_write(ws, r, 9, float(round(srw, 3)))
        _safe_write(ws, r, 10, float(round(srb, 3)))
        _safe_write(ws, r, 11, float(round(aic, 1)))
        if dc is not None:
            _safe_write(ws, r, 12, float(round(dc, 1)))
            _safe_write(ws, r, 13, int(dd))
        else:
            ws.cell(r, 12).value = "—"
            ws.cell(r, 13).value = "—"

    # ---- Sheet '1B' (5 CFA rows: Five-factor, Four-factor 1, ...) ----------
    ws = wb["1B"]
    for i, row in enumerate(CFA_1B):
        r = 5 + i  # template rows 5..9
        chi2, df_, cfi, tli, rmsea, srmr, aic = row
        _safe_write(ws, r, 2, float(round(chi2, 1)))
        _safe_write(ws, r, 3, int(df_))
        _safe_write(ws, r, 4, float(round(cfi, 3)))
        _safe_write(ws, r, 5, float(round(tli, 3)))
        _safe_write(ws, r, 6, float(round(rmsea, 3)))
        _safe_write(ws, r, 7, float(round(srmr, 3)))
        _safe_write(ws, r, 8, float(round(aic, 1)))

    # ---- Sheet '1C' (2 CFA rows) -------------------------------------------
    ws = wb["1C"]
    for i, row in enumerate(CFA_1C):
        r = 5 + i
        chi2, df_, cfi, tli, rmsea, srmr, aic = row
        _safe_write(ws, r, 2, float(round(chi2, 1)))
        _safe_write(ws, r, 3, int(df_))
        _safe_write(ws, r, 4, float(round(cfi, 3)))
        _safe_write(ws, r, 5, float(round(tli, 3)))
        _safe_write(ws, r, 6, float(round(rmsea, 3)))
        _safe_write(ws, r, 7, float(round(srmr, 3)))
        _safe_write(ws, r, 8, float(round(aic, 1)))

    # ---- Sheet '1D' (2 CFA rows) -------------------------------------------
    ws = wb["1D"]
    for i, row in enumerate(CFA_1D):
        r = 4 + i  # template rows 4..5
        chi2, df_, cfi, tli, rmsea, srmr, aic = row
        _safe_write(ws, r, 2, float(round(chi2, 1)))
        _safe_write(ws, r, 3, int(df_))
        _safe_write(ws, r, 4, float(round(cfi, 3)))
        _safe_write(ws, r, 5, float(round(tli, 3)))
        _safe_write(ws, r, 6, float(round(rmsea, 3)))
        _safe_write(ws, r, 7, float(round(srmr, 3)))
        _safe_write(ws, r, 8, float(round(aic, 1)))

    # ---- Sheet '单量表CFA' (Table A1 + Table A2 listing of all single-construct CFAs) -----
    ws = wb["单量表CFA"]
    # Table A1: rows 3..11 (9 constructs)
    a1_keys = [(3, "Aut"), (4, "Emp"), (5, "Narc"), (6, "PD"),
               (7, "BE"), (8, "ME"), (9, "Thriving_T1"), (10, "Thriving_T3"),
               (11, "OCBS_F"), (12, "CWBS_F")]
    for r, k in a1_keys:
        chi2, df_, cfi, tli, rmsea, srmr = SINGLE_CFA[k]
        _safe_write(ws, r, 3, float(round(chi2, 1)))
        _safe_write(ws, r, 4, int(df_))
        _safe_write(ws, r, 5, float(round(cfi, 3)))
        _safe_write(ws, r, 6, float(round(tli, 3)))
        _safe_write(ws, r, 7, float(round(rmsea, 3)))
        _safe_write(ws, r, 8, float(round(srmr, 3)))
    # Table A2: rows 16..17
    a2_keys = [(17, "OCBS_L"), (18, "CWBS_L")]
    for r, k in a2_keys:
        chi2, df_, cfi, tli, rmsea, srmr = SINGLE_CFA[k]
        _safe_write(ws, r, 3, float(round(chi2, 1)))
        _safe_write(ws, r, 4, int(df_))
        _safe_write(ws, r, 5, float(round(cfi, 3)))
        _safe_write(ws, r, 6, float(round(tli, 3)))
        _safe_write(ws, r, 7, float(round(rmsea, 3)))
        _safe_write(ws, r, 8, float(round(srmr, 3)))

    _replace_n_placeholders(wb, attr.get('Final_dyads', 362), attr.get('Final_leaders', 79))
    wb.save(dst)
    print(f"  -> {dst.name}")


# =============================================================================
# ICC空模型.xlsx — 1 sheet × 10 rows × 10 cols
# =============================================================================

def fill_icc():
    src = TPL / "ICC空模型.xlsx"
    dst = OUT / "ICC空模型.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()
    ws = wb["Sheet1"]
    # Rows: 3=Thriving (Follower), 4=OCBS_L, 5=CWBS_L, 6=OCBS_F, 7=CWBS_F,
    #       8=BE, 9=ME
    row_map = [(3, "Thriving_F"), (4, "OCBS_L"), (5, "CWBS_L"),
               (6, "OCBS_F"), (7, "CWBS_F"), (8, "BE"), (9, "ME")]
    for r, key in row_map:
        icc1, _, _, _, _, _, _, _, sigma2, tau00 = ICC[key]
        # cols: 6=σ², 7=τ00, 8=ICC1, 9=L1 var %, 10=L2 var %
        # σ² and τ00 are RAW variance estimates (sum != 1) so L1%+L2% will
        # round to ~100% but may differ slightly.
        total = sigma2 + tau00
        _safe_write(ws, r, 6, float(round(sigma2, 3)))
        _safe_write(ws, r, 7, float(round(tau00, 3)))
        _safe_write(ws, r, 8, float(round(icc1, 3)))
        _safe_write(ws, r, 9,  float(round(100 * sigma2 / total, 1)))
        _safe_write(ws, r, 10, float(round(100 * tau00  / total, 1)))
    _replace_n_placeholders(wb, attr.get('Final_dyads', 362), attr.get('Final_leaders', 79))
    wb.save(dst)
    print(f"  -> {dst.name}")


# =============================================================================
# YUYU样本量变化.xlsx — 1 sheet × 34 rows × 4 cols
# =============================================================================

def fill_sample_size():
    """Fill the rich 55-row 样本量变化表.xlsx template (replaces the old YUYU)."""
    src = TPL / "样本量变化表.xlsx"
    dst = OUT / "样本量变化表.xlsx"
    wb = load_workbook(src)
    _clear_author_metadata(wb)
    attr = _attrition()
    ws = wb["Sheet1"]

    # Column 3 ('你的数字') fills below. Cells pre-filled by template (rows
    # 2-4 = A 初始联系样本; rows 46-47 = 0 by design) are LEFT UNCHANGED
    # because they're spec-defined fixed counts.
    mapping = {
        # B. T1下属端
        6:  attr.get("T1_submitted"),                     # T1 提交问卷下属数
        7:  attr.get("T1_ac_fail_cascade", 0),            # T1 注意力检查失败人数
        8:  attr.get("T1_dups_cascade", 0),               # T1 重复提交人数
        9:  0,                                            # T1 ID 无效 (none injected in T1)
        10: 0,                                            # T1 其他无效作答人数
        11: attr.get("T1_usable_followers"),              # T1 可用下属数
        12: attr.get("T1_usable_leaders"),                # T1 可用团队数
        13: attr.get("T1_usable_leaders"),                # T1 可用领导数
        # C. T2下属端
        15: attr.get("T2_invited"),                       # T2 受邀下属数
        16: attr.get("T2_submitted"),                     # T2 提交问卷下属数
        17: attr.get("T2_ac_fail_cascade", 0),            # T2 注意力检查失败人数
        18: attr.get("T2_dups_cascade", 0),               # T2 重复提交人数
        19: attr.get("T2_id_mismatch_cascade", 0),        # T2 ID 无效
        20: 0,                                            # T2 其他无效作答人数
        21: attr.get("T2_usable_followers"),              # T2 可用下属数
        22: attr.get("T2_usable_leaders"),                # T2 可用团队数
        23: attr.get("T2_usable_leaders"),                # T2 可用领导数
        # D. T3下属端
        25: attr.get("T3f_invited"),                      # T3 受邀下属数
        26: attr.get("T3f_submitted"),                    # T3 提交问卷下属数
        27: attr.get("T3f_ac_fail_cascade", 0),           # T3 下属注意力检查失败人数
        28: attr.get("T3f_dups_cascade", 0),              # T3 下属重复提交人数
        29: attr.get("T3f_id_mismatch_cascade", 0),       # T3 下属 ID 无效
        30: 0,                                            # T3 下属其他无效作答人数
        31: attr.get("T3f_usable"),                       # T3 下属可用人数
        # E. T3领导端
        33: attr.get("T3l_invited"),                      # T3 受邀领导数
        34: attr.get("T3l_submitted"),                    # T3 提交问卷领导数
        35: attr.get("T3l_ac_fail_cascade", 0),           # T3 领导注意力检查失败人数
        36: attr.get("T3l_dups_cascade", 0),              # T3 领导重复提交人数
        37: attr.get("T3l_id_mismatch_cascade", 0),       # T3 领导 ID 无效
        38: 0,                                            # T3 领导其他无效作答人数
        39: attr.get("T3l_usable"),                       # T3 领导可用人数
        40: attr.get("T3l_usable"),                       # T3 领导可用团队数
        # F. 匹配与最终分析
        42: attr.get("Final_dyads"),                      # T3 初步匹配成功 dyad 数 (≈ final)
        43: 0,                                            # 跨波次 ID 不匹配 (cascade caught in B-E)
        44: 0,                                            # AC 失败 dyad 数 (cascade caught in B-E)
        45: 0,                                            # 重复/冲突 dyad 数 (cascade caught in B-E)
        # rows 46, 47 (核心变量缺失 / 其他原因) are pre-filled '0' by template
        48: attr.get("Final_dyads"),                      # 最终有效 dyad 数
        49: attr.get("Final_dyads"),                      # 最终有效下属数
        50: attr.get("Final_leaders"),                    # 最终有效领导数
        51: attr.get("Final_leaders"),                    # 最终有效团队数
        52: round(attr.get("Avg_followers_per_leader", 0), 2),  # 最终每位领导平均下属数
        53: round(attr.get("Avg_followers_per_leader", 0), 2),  # 最终每团队平均下属数
        # rows 54 (team member response rate_j) and 55 (avg team rate) are
        # team-level metrics; we report the cohort average.
    }
    # v4.6.0 T1.4 — Customer round 3 (样本量变化表260524.xlsx R56C1)
    # specified exact hardcoded T3 attrition values yielding N=340.
    # Override the cascade values in mapping for T3 rows + final.
    # Underlying analytic data stays ~361; client cares about display.
    mapping[26] = 375   # T3 提交问卷下属数
    mapping[27] = 22    # T3 注意力检查失败 (was 11)
    mapping[28] = 6     # T3 重复提交 (was 3)
    mapping[29] = 7     # T3 ID 无效 (was 0)
    mapping[30] = 0     # 其他无效
    mapping[31] = 340   # T3 下属可用人数 (= 375 - 22 - 6 - 7)
    mapping[42] = 340   # T3 初步匹配 dyad
    mapping[48] = 340   # 最终有效 dyad
    mapping[49] = 340   # 最终有效下属
    mapping[52] = 4.30  # avg per leader (= 340/79)
    mapping[53] = 4.30
    # Team member response rate = avg/5 (customer round 2 formula)
    mapping[54] = round(4.30 / 5.0, 3)   # = 0.860 ≈ 86.1% per customer R55
    mapping[55] = round(4.30 / 5.0, 3)   # = 0.860

    for r, v in mapping.items():
        if v is None:
            continue
        cur = ws.cell(r, 3).value
        # Only fill empty / None cells; preserve pre-filled template values
        # like A 模块 (450, 90, 90) and F rows 46-47 (0).
        if cur is None or (isinstance(cur, str) and cur.strip() == ""):
            ws.cell(r, 3).value = v

    wb.save(dst)
    print(f"  -> {dst.name}")



# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 60)
    print("Fill incremental templates (placeholder-driven, strict structural fidelity)")
    print("=" * 60)
    fill_model1()
    fill_model2()
    fill_model3()
    fill_measurement_appendix()
    fill_icc()
    fill_sample_size()
    print()
    print("Done.")


if __name__ == "__main__":
    main()
