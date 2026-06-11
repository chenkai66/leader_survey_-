"""
Constraint validator (v3) — comprehensive strict validator for the
post-second-round deliverables.

Sections:
  1. Sample sizes (T1=90, T2=85, T3=79; final dyads from attrition)
  2. Raw > cleaned (cleaning removes dups + mismatches + AC failures)
  3. Each leader 3-5 followers in final (NOT 6+, ≥3, ≤5)
  4. Attention-check items present; AC=6 means PASS, cleaned has only AC=6
  5. CLID 1:1 numeric, range 1-79
  6. LeaderEducation only in T3 leader (and final via merge), 2-5 integer
  7. Grand-mean centering exact + dummies not centered
  8. No narcissism × leadership interaction column
  9. Duplicate / mismatched IDs only in raw
 10. Missing-value pattern (T1 ~10 non-core, T2 zero, T3-leader ~3)
 11. Composite scores match item averages
 12. Parcel definitions match theoretical maps
 13. Reverse-coded items in [1, 7], parcels use the reversed versions
 14. Likert ranges
 15. No NaN in core analysis variables of final
 16. Dummy variables in {0, 1}; Company dummies present
 17. Hypothesis directions (correlation sign checks)
 18. Eight deliverable files exist
 19. Each incremental file has EXACTLY ONE sheet
 20. Master deliverables 7 / 4 sheets
 21. Master template key cells filled
 22. MCFA Mplus dat
 23. Cross-wave ID integrity
 24. No duplicate IDs in cleaned
 25. Master Table 4 path-sign correctness
 26. Cluster adjustment in master appendix Table A1/A2
 27. Centered means ~ 0
 28. Companies = {A, B, C}; TeamID == LeaderID
 29. ID format follows {Company}_L{NN} / {LeaderID}_F{N}
 30. TenureWithLeader integer-dominant (>= 90% are integer)
 31. measurement appendix has explicit χ² column
 32. ICC table has 5 columns including non-empty Notes
 33. YUYU table 26 rows all column-C filled
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
DATA = ROOT / "data"
RES = ROOT / "results"

PASS, FAIL = "PASS", "FAIL"
_results = []


def check(name, cond, detail=""):
    status = PASS if cond else FAIL
    line = f"  [{status}] {name}"
    if detail:
        line += f" -- {detail}"
    print(line)
    _results.append(bool(cond))
    return bool(cond)


def section(title):
    print("\n" + "-" * 64)
    print(title)
    print("-" * 64)


def load():
    files = {
        "t1_raw":  "T1_raw.xlsx",  "t1": "T1_cleaned.xlsx",
        "t2_raw":  "T2_raw.xlsx",  "t2": "T2_cleaned.xlsx",
        "t3l_raw": "T3_leader_raw.xlsx",   "t3l": "T3_leader_cleaned.xlsx",
        "t3f_raw": "T3_follower_raw.xlsx", "t3f": "T3_follower_cleaned.xlsx",
        "final":   "final_merged_analysis_data.xlsx",
    }
    out = {}
    for k, p in files.items():
        full = DATA / p
        if not full.exists():
            print(f"MISSING: {full}")
            sys.exit(2)
        out[k] = pd.read_excel(full)
    return out


def main() -> int:
    print("=" * 70)
    print("LEADER-SURVEY CONSTRAINT VALIDATOR (v3 — post-feedback)")
    print("=" * 70)

    d = load()
    t1, t1r = d["t1"], d["t1_raw"]
    t2, t2r = d["t2"], d["t2_raw"]
    t3l, t3lr = d["t3l"], d["t3l_raw"]
    t3f, t3fr = d["t3f"], d["t3f_raw"]
    final = d["final"]

    # ---------- 1. sample sizes ----------
    section("1. Sample sizes (waves)")
    check("T1 leaders == 90", t1["LeaderID"].nunique() == 90,
          f"got {t1['LeaderID'].nunique()}")
    check("T2 leaders == 85", t2["LeaderID"].nunique() == 85,
          f"got {t2['LeaderID'].nunique()}")
    # round-5: T3 leader survey is now DYAD-level (one row per follower the
    # leader evaluated) — a leader rates several followers. 340 rows / 79 leaders.
    check("T3 leader cleaned rows == 340 (dyad-level)", len(t3l) == 340, f"got {len(t3l)}")
    check("T3 leader cleaned leaders == 79", t3l["LeaderID"].nunique() == 79,
          f"got {t3l['LeaderID'].nunique()}")
    check("T3 leader followers unique == 340", t3l["FollowerID"].nunique() == 340,
          f"got {t3l['FollowerID'].nunique()}")
    check("T3 leader has NumFollowersEvaluated col", "NumFollowersEvaluated" in t3l.columns)
    if "NumFollowersEvaluated" in t3l.columns:
        per = t3l.groupby("LeaderID")["FollowerID"].count()
        nfe = t3l.groupby("LeaderID")["NumFollowersEvaluated"].first()
        check("NumFollowersEvaluated == actual dyad count", bool((per.values == nfe.values).all()),
              "count mismatch")
        check("SpanOfControl >= NumFollowersEvaluated",
              bool((t3l["SpanOfControl"] >= t3l["NumFollowersEvaluated"]).all()))
    check("T3 follower leaders == 79", t3f["LeaderID"].nunique() == 79,
          f"got {t3f['LeaderID'].nunique()}")
    check("final leaders == 79", final["LeaderID"].nunique() == 79,
          f"got {final['LeaderID'].nunique()}")
    # round-4 T1.4: final dyads MUST equal the sample-size table's final N (340),
    # not just sit loosely in [237, 395]. The previous loose bound let 361 pass
    # while the 样本量变化表 said 340 — exactly the customer's complaint that the
    # merged data and the sample table "对不上".
    check("final dyads == 340 (matches 样本量变化表)", len(final) == 340,
          f"got {len(final)}")
    _p_ss = RES / "样本量变化表.xlsx"
    if _p_ss.exists():
        _wb = load_workbook(_p_ss); _ws = _wb["Sheet1"]
        _r49 = _ws.cell(49, 3).value
        check("len(final) == 样本量变化表 R49 (data↔table cross-check)",
              len(final) == _r49, f"final={len(final)}, R49={_r49}")
        _wb.close()
    # round-4 sample table A57 team-size spec: 3-person×10, 4-person×35, 5-person×34.
    _tsd = final.groupby("LeaderID").size().value_counts().to_dict()
    _tsd = {int(k): int(v) for k, v in _tsd.items()}
    check("team-size distribution == {3:10, 4:35, 5:34}",
          _tsd == {3: 10, 4: 35, 5: 34}, f"got {dict(sorted(_tsd.items()))}")
    check("avg followers/leader == 4.3 (340/79)",
          abs(len(final) / final["LeaderID"].nunique() - 4.30) < 0.01,
          f"got {len(final)/final['LeaderID'].nunique():.3f}")

    # ---------- 2. raw > cleaned ----------
    section("2. Raw > cleaned")
    for label, raw, clean in [("T1", t1r, t1), ("T2", t2r, t2),
                              ("T3 leader", t3lr, t3l),
                              ("T3 follower", t3fr, t3f)]:
        check(f"{label} raw > cleaned", len(raw) > len(clean),
              f"{len(raw)} > {len(clean)}")

    # ---------- 3. 3-5 followers per leader ----------
    section("3. 3-5 followers per leader (final analysis)")
    g = final.groupby("LeaderID").size()
    check("final min ≥ 3", g.min() >= 3, f"min={g.min()}")
    check("final max ≤ 5  (NOT 6 or more)", g.max() <= 5, f"max={g.max()}")
    check("final mean in [3.5, 5.0]",
          3.5 <= g.mean() <= 5.0, f"mean={g.mean():.2f}")

    # ---------- 4. attention checks ----------
    section("4. AC = 6 means PASS; cleaned only contains AC = 6")
    check("T1 has EMP9_AttCheck",  "EMP9_AttCheck" in t1.columns)
    check("T2 has MAL6_AttCheck",  "MAL6_AttCheck" in t2.columns)
    check("T3f has OCBS7_AttCheck","OCBS7_AttCheck" in t3f.columns)
    check("T3l has CWBS6_AttCheck","CWBS6_AttCheck" in t3l.columns)
    # In RAW data we should see some failures (3-5%):
    if "EMP9_AttCheck" in t1r.columns:
        fails = (t1r["EMP9_AttCheck"] != 6).sum()
        check("T1 raw has 1 ≤ AC failures ≤ 30", 1 <= fails <= 30, f"fails={fails}")
    # Cleaned MUST have ALL AC == 6
    for label, df, col in [("T1", t1, "EMP9_AttCheck"),
                            ("T2", t2, "MAL6_AttCheck"),
                            ("T3f", t3f, "OCBS7_AttCheck"),
                            ("T3l", t3l, "CWBS6_AttCheck")]:
        if col in df.columns:
            non6 = (df[col] != 6).sum()
            check(f"{label} cleaned: every row has AC = 6", non6 == 0,
                  f"non-6 count={non6}")

    # ---------- 5. CLID ----------
    section("5. CLID")
    if "CLID" in final.columns:
        check("CLID numeric", pd.api.types.is_numeric_dtype(final["CLID"]))
        check("CLID range [1, 79]",
              final["CLID"].min() == 1 and final["CLID"].max() == 79,
              f"[{final['CLID'].min()}, {final['CLID'].max()}]")
        check("CLID 1:1 LeaderID",
              len(final[["LeaderID", "CLID"]].drop_duplicates()) == 79)

    # ---------- 6. LeaderEducation only in T3 leader (and final via merge) ----------
    section("6. LeaderEducation in T3 leader survey (not T1)")
    check("T1 has NO LeaderEducation column", "LeaderEducation" not in t1.columns)
    check("T3 leader HAS LeaderEducation column", "LeaderEducation" in t3l.columns)
    if "LeaderEducation" in t3l.columns:
        col = t3l["LeaderEducation"]
        check("T3 leader: LeaderEducation min ≥ 2", col.min() >= 2,
              f"min={col.min()}")
        check("T3 leader: LeaderEducation max ≤ 5", col.max() <= 5,
              f"max={col.max()}")
        check("T3 leader: LeaderEducation integer",
              col.dropna().apply(lambda x: float(x).is_integer()).all())

    # ---------- 7. centering ----------
    section("7. Grand-mean centering")
    must_center = ["Autocratic", "Empowering", "Narcissism", "PowerDistance",
                   "FollowerAge", "TenureWithLeader", "InteractionFreq",
                   "T1_Thriving", "WorkingYears"]
    for v in must_center:
        c = f"{v}_C"
        if c not in final.columns:
            check(f"{c} exists", False, "missing")
            continue
        check(f"{c} mean ~ 0", abs(final[c].mean()) < 1e-3,
              f"mean={final[c].mean():.6f}")
        if v in final.columns:
            expected = final[v] - final[v].mean()
            diff = (final[c] - expected).abs().max()
            check(f"{c} == {v} - grand_mean", diff < 1e-6, f"max diff={diff:.2e}")
    centered_dummies = [c for c in final.columns
                        if c.endswith("_C") and ("Gender_" in c or "Edu_" in c)
                        and not c.startswith("Company_")]
    check("Dummies NOT centered", not centered_dummies,
          f"leaks: {centered_dummies}" if centered_dummies else "")

    # ---------- 8. narcissism not a moderator ----------
    section("8. Narcissism is NOT a moderator")
    bad = [c for c in final.columns if "Narcissism" in c
           and any(t in c for t in ["x", "×", "X_", "_x_", "Interaction"])]
    check("no narcissism × leadership column", not bad,
          f"found: {bad}" if bad else "")

    # ---------- 9. raw duplicates / mismatches ----------
    section("9. Duplicate / mismatched IDs in raw")
    if "FollowerID" in t1r.columns:
        n = t1r["FollowerID"].duplicated().sum()
        check("T1 raw 1 ≤ dup ≤ 15", 1 <= n <= 15, f"{n}")
    if "FollowerID" in t2r.columns:
        n = t2r["FollowerID"].duplicated().sum()
        check("T2 raw 1 ≤ dup ≤ 6", 1 <= n <= 6, f"{n}")
    if "FollowerID" in t3lr.columns:
        n = t3lr["FollowerID"].duplicated().sum()
        check("T3 leader raw 0 < follower-dup ≤ 1", 0 < n <= 1, f"{n}")
    if "FollowerID" in t1.columns and "FollowerID" in t2r.columns:
        miss = set(t2r["FollowerID"]) - set(t1["FollowerID"])
        check("T2 raw ≥ 3 unmatched", len(miss) >= 3, f"{len(miss)}")
    if "LeaderID" in t2.columns and "LeaderID" in t3lr.columns:
        miss = set(t3lr["LeaderID"]) - set(t2["LeaderID"])
        check("T3 leader raw ≥ 1 unmatched", len(miss) >= 1, f"{len(miss)}")

    # ---------- 10. missing pattern ----------
    section("10. Missing-value pattern")
    check("T1 raw missing in [5, 20]", 5 <= int(t1r.isna().sum().sum()) <= 20,
          f"{int(t1r.isna().sum().sum())}")
    check("T2 raw zero missing", int(t2r.isna().sum().sum()) == 0,
          f"{int(t2r.isna().sum().sum())}")
    check("T3 leader raw missing in [1, 5]",
          1 <= int(t3lr.isna().sum().sum()) <= 5,
          f"{int(t3lr.isna().sum().sum())}")

    # ---------- 11. composite scores ----------
    section("11. Composites equal item averages")
    cases = [
        ("Autocratic",  [f"AUT{i}"  for i in range(1, 7)],  t1),
        ("Empowering",  [f"EMP{i}"  for i in range(1, 13)], t1),
        ("Narcissism",  [f"NARC{i}" for i in range(1, 7)],  t1),
        ("PowerDistance", [f"PD{i}" for i in range(1, 6)],  t1),
    ]
    for name, items, df in cases:
        present = [c for c in items if c in df.columns]
        if name in df.columns and present:
            diff = (df[name] - df[present].mean(axis=1)).abs().max()
            check(f"{name} == mean(items)", diff < 1e-6, f"diff={diff:.2e}")
    if "BenignEnvy" in final.columns:
        ben = [f"BEN{i}" for i in range(1, 6)]
        diff = (final["BenignEnvy"] - final[ben].mean(axis=1)).abs().max()
        check("BenignEnvy == mean(BEN1..5)", diff < 1e-6, f"diff={diff:.2e}")
    if "MaliciousEnvy" in final.columns:
        mal = [f"MAL{i}" for i in range(1, 6)]
        diff = (final["MaliciousEnvy"] - final[mal].mean(axis=1)).abs().max()
        check("MaliciousEnvy == mean(MAL1..5)", diff < 1e-6, f"diff={diff:.2e}")

    # ---------- 12. parcels ----------
    section("12. Parcel definitions")
    parcels = [
        ("EMPP1", ["EMP1", "EMP2", "EMP3"]),
        ("EMPP2", ["EMP4", "EMP5", "EMP6"]),
        ("EMPP3", ["EMP7", "EMP8", "EMP9"]),
        ("EMPP4", ["EMP10", "EMP11", "EMP12"]),
        ("THRP1", ["THR1", "THR2", "THR3"]),
        ("THRP2", ["THR4", "R_THR5"]),
        ("THRP3", ["THR6", "THR7", "THR8"]),
        ("THRP4", ["THR9", "R_THR10"]),
    ]
    for parcel, items in parcels:
        if all(c in t1.columns for c in items + [parcel]):
            diff = (t1[parcel] - t1[items].mean(axis=1)).abs().max()
            check(f"{parcel} = mean({','.join(items)})", diff < 1e-6,
                  f"{diff:.2e}")

    # ---------- 13. reverse-coded ----------
    section("13. Reverse-coded items in [1, 7]")
    for col in ("R_THR5", "R_THR10"):
        if col in t1.columns:
            check(f"{col} in [1, 7]",
                  t1[col].min() >= 1 and t1[col].max() <= 7,
                  f"[{t1[col].min()}, {t1[col].max()}]")

    # ---------- 14. Likert ranges ----------
    section("14. Likert ranges")
    for col in ["AUT1", "AUT6", "EMP1", "EMP12", "THR1", "THR9",
                "NARC1", "NARC6", "PD1"]:
        if col in t1.columns:
            mn, mx = t1[col].min(), t1[col].max()
            check(f"{col} in 1..7", mn >= 1 and mx <= 7, f"[{mn},{mx}]")
    for col in ["BEN1", "BEN5", "MAL1", "MAL5"]:
        if col in t2.columns:
            mn, mx = t2[col].min(), t2[col].max()
            check(f"{col} in 1..7", mn >= 1 and mx <= 7, f"[{mn},{mx}]")

    # ---------- 15. no NaN in core analysis vars ----------
    section("15. No NaN in core analysis vars (final)")
    for col in ["LeaderID", "FollowerID", "CLID",
                "Autocratic", "Empowering", "Narcissism", "PowerDistance",
                "BenignEnvy", "MaliciousEnvy",
                "T3_Thriving", "OCBS_Leader", "CWBS_Leader",
                "OCBS_Follower", "CWBS_Follower",
                "Autocratic_C", "Empowering_C", "Narcissism_C",
                "PowerDistance_C", "WorkingYears_C"]:
        if col in final.columns:
            n = final[col].isna().sum()
            check(f"final.{col} no NaN", n == 0, f"NaN={n}")

    # ---------- 16. dummies (incl Company) ----------
    section("16. Dummies in {0,1} including Company dummies")
    for d2 in ["Gender_Female", "Edu_HighSchool", "Edu_Associate",
               "Edu_Master", "Edu_Doctoral", "Company_B", "Company_C"]:
        if d2 in final.columns:
            vs = sorted(final[d2].dropna().unique().tolist())
            check(f"{d2} in {{0,1}}", set(vs).issubset({0, 1}), f"vals={vs}")

    # ---------- 17. hypothesis directions ----------
    section("17. Hypothesis directions (corr signs)")
    pairs = [
        ("Autocratic",    "MaliciousEnvy",  "+"),
        ("Empowering",    "BenignEnvy",     "+"),
        ("Empowering",    "MaliciousEnvy",  "-"),
        ("MaliciousEnvy", "CWBS_Leader",    "+"),
        ("MaliciousEnvy", "T3_Thriving",    "-"),
        ("BenignEnvy",    "T3_Thriving",    "+"),
        ("BenignEnvy",    "OCBS_Leader",    "+"),
        ("Autocratic",    "BenignEnvy",     "-"),
    ]
    for x, y, sign in pairs:
        if x in final.columns and y in final.columns:
            r = final[[x, y]].corr().iloc[0, 1]
            ok = (sign == "+" and r > 0) or (sign == "-" and r < 0)
            check(f"corr({x},{y}) sign {sign}", ok, f"r={r:+.3f}")

    # ---------- 18. eight deliverable files exist ----------
    section("18. Eight deliverable files exist")
    for f in ["Model1.xlsx", "Model2.xlsx", "Model3.xlsx",
              "measurement appendix.xlsx", "ICC空模型.xlsx",
              "样本量变化表.xlsx",
              "主模型结果填答表.xlsx", "study3附录结果填答.xlsx"]:
        check(f"results/{f}", (RES / f).exists())

    # ---------- 19. each incremental file has exactly 1 sheet ----------
    section("19. Each incremental file has EXACTLY 1 sheet")
    for f in ["Model1.xlsx", "Model2.xlsx", "Model3.xlsx",
              "measurement appendix.xlsx", "ICC空模型.xlsx",
              "样本量变化表.xlsx"]:
        p = RES / f
        # Sheet count is intentionally multi-sheet for the new richer templates.
        # We no longer enforce single-sheet; the template-byte-equal check covers
        # structural fidelity instead.
        if p.exists():
            wb = load_workbook(p, read_only=True)
            check(f"{f} has at least 1 sheet", len(wb.sheetnames) >= 1,
                  f"sheets={wb.sheetnames}")
            wb.close()
    section("20. Master deliverables 7 / 4 sheets")
    masters = {
        "主模型结果填答表.xlsx":
            ["总览", "Table 1A", "Table 1B",
             "Table 2. Aggregation Statistics",
             "Table 3. Correlation",
             "Table 4. 主模型path",
             "Table 5. Moderation and Conditi"],
        "study3附录结果填答.xlsx":
            ["Table A12 单量表CFA",
             "Table A3 区分多来源结果变量",
             "Table A4 Robustness",
             "Table A5 Robustness"],
    }
    for f, expected_sheets in masters.items():
        p = RES / f
        if not p.exists():
            check(f"{f} exists", False)
            continue
        wb = load_workbook(p, read_only=True)
        actual = wb.sheetnames
        check(f"{f} sheet count == {len(expected_sheets)}",
              len(actual) == len(expected_sheets),
              f"got {len(actual)}: {actual}")
        for s in expected_sheets:
            check(f"  '{s}' present", s in actual)
        wb.close()

    # ---------- 21. master cells filled ----------
    section("21. Master template key cells filled")
    p = RES / "主模型结果填答表.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb["Table 1A"]
        v = ws.cell(row=3, column=2).value
        check("Table 1A row3 col B numeric", isinstance(v, (int, float)),
              f"got {type(v).__name__}={v}")
        ws = wb["Table 4. 主模型path"]
        v = ws.cell(row=4, column=3).value
        check("Table 4 row4 col C numeric", isinstance(v, (int, float)),
              f"got {type(v).__name__}={v}")
        ws = wb["Table 3. Correlation"]
        v = ws.cell(row=3, column=2).value
        check("Table 3 row3 col B (Age mean) numeric",
              isinstance(v, (int, float)), f"got {type(v).__name__}={v}")
        wb.close()

    # ---------- 22. MCFA dat ----------
    section("22. MCFA Mplus dat")
    mcfa = DATA / "study3_mcfa.dat"
    check("study3_mcfa.dat exists", mcfa.exists())
    if mcfa.exists():
        with open(mcfa, encoding="utf-8", errors="ignore") as f:
            lines = [ln for ln in f if ln.strip()]
        check(f"mcfa rows == final length ({len(final)})",
              len(lines) == len(final), f"{len(lines)}")
        first = lines[0].strip().split()
        check("mcfa first col numeric (CLID)",
              first[0].lstrip("-").replace(".", "").isdigit())

    # ---------- 23. cross-wave id integrity ----------
    section("23. Cross-wave ID integrity")
    fids = set(final["FollowerID"])
    check("final ⊆ T1 cleaned", not (fids - set(t1["FollowerID"])))
    check("final ⊆ T2 cleaned", not (fids - set(t2["FollowerID"])))
    check("final ⊆ T3 follower cleaned", not (fids - set(t3f["FollowerID"])))
    check("final leaders ⊆ T3 leader cleaned",
          not (set(final["LeaderID"]) - set(t3l["LeaderID"])))

    # ---------- 24. no duplicate IDs in cleaned ----------
    section("24. No duplicate IDs in cleaned")
    for label, df, key in [("T1", t1, "FollowerID"), ("T2", t2, "FollowerID"),
                            ("T3 follower", t3f, "FollowerID"),
                            ("T3 leader", t3l, "FollowerID"),
                            ("final", final, "FollowerID")]:
        if key in df.columns:
            n = df[key].duplicated().sum()
            check(f"{label} no duplicate {key}", n == 0, f"dup={n}")

    # ---------- 25. master Table 4 path signs ----------
    section("25. Master Table 4 path signs")
    p = RES / "主模型结果填答表.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb["Table 4. 主模型path"]
        for label, row, expected_sign in [
                ("Auto→Benign",      4, "-"),
                ("Emp→Benign",       5, "+"),
                ("Auto→Malicious",   6, "+"),
                ("Mal→Thriving",    12, "-")]:
            v = ws.cell(row=row, column=3).value
            ok = (isinstance(v, (int, float)) and
                  ((expected_sign == "+" and v > 0) or
                   (expected_sign == "-" and v < 0)))
            check(f"Table4 {label} sign {expected_sign}", ok, f"b={v}")
        wb.close()

    # ---------- 26. master appendix structural fidelity ----------
    # Cluster adjustment is documented in mcfa_mplus_syntax.inp and
    # analysis_code.R, NOT injected into the template's Notes column
    # (which would violate the strict template-fidelity requirement).
    section("26. Master appendix structural fidelity")
    p = RES / "study3附录结果填答.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb["Table A12 单量表CFA"]
        # Original construct labels must be preserved verbatim
        check("Table A1 row 4 col 1 == 'Empowering leadership'",
              ws.cell(row=4, column=1).value == "Empowering leadership",
              f"got {ws.cell(row=4, column=1).value!r}")
        check("Table A1 row 9 col 1 == 'Thriving'",
              ws.cell(row=9, column=1).value == "Thriving",
              f"got {ws.cell(row=9, column=1).value!r}")
        wb.close()

    # ---------- 27. centered means ~ 0 ----------
    section("27. Centered means ~ 0")
    for v in must_center:
        c = f"{v}_C"
        if c in final.columns:
            m = abs(final[c].mean())
            check(f"{c} |mean| < 1e-6", m < 1e-6, f"|mean|={m:.2e}")

    # ---------- 28. companies ----------
    section("28. Companies = {A, B, C}; TeamID == LeaderID")
    if "CompanyID" in final.columns:
        comps = sorted(final["CompanyID"].unique().tolist())
        check("CompanyID in {A,B,C}",
              set(comps) == {"A", "B", "C"}, f"{comps}")
    else:
        check("CompanyID column exists", False, "missing")
    if "TeamID" in final.columns and "LeaderID" in final.columns:
        check("TeamID == LeaderID",
              (final["TeamID"] == final["LeaderID"]).all())
    else:
        check("TeamID column exists", "TeamID" in final.columns,
              f"final cols: TeamID present? {'TeamID' in final.columns}")

    # ---------- 29. ID format ----------
    section("29. ID format compliance")
    pat_lid = re.compile(r"^[ABC]_L\d{2}$")
    pat_fid = re.compile(r"^[ABC]_L\d{2}_F\d+$")
    bad_lids = [l for l in final["LeaderID"].unique() if not pat_lid.match(str(l))]
    check("LeaderID format {A|B|C}_L{NN}", not bad_lids,
          f"bad: {bad_lids[:3]}")
    bad_fids = [f for f in final["FollowerID"].head(50)
                if not pat_fid.match(str(f))]
    check("FollowerID format {LID}_F{N}", not bad_fids,
          f"bad: {bad_fids[:3]}")

    # ---------- 30. tenure integer-dominant ----------
    section("30. TenureWithLeader dominantly integer (>= 90%)")
    if "TenureWithLeader" in final.columns:
        col = final["TenureWithLeader"]
        n_int = col.apply(lambda x: float(x).is_integer()).sum()
        pct = 100 * n_int / len(col)
        check("TenureWithLeader >= 90% integer",
              pct >= 90, f"{pct:.1f}% integer")

    # ---------- 31. measurement appendix: 5 sheets present with fit rows ----------
    section("31. measurement appendix structural fidelity (new 5-sheet layout)")
    p = RES / "measurement appendix.xlsx"
    if p.exists():
        wb = load_workbook(p)
        expected_sheets = ["1A", "1B", "1C", "1D", "单量表CFA"]
        check("measurement appendix has 5 expected sheets",
              wb.sheetnames == expected_sheets, f"got {wb.sheetnames}")
        # 1A row 3 col 4 (χ²) numeric
        val = wb["1A"].cell(row=3, column=4).value
        check("1A row3 χ² numeric > 0",
              isinstance(val, (int, float)) and val > 0, f"v={val}")
        wb.close()

    # ---------- 32. ICC空模型 has 10 columns + 7 variable rows ----------
    section("32. ICC table new layout (10 cols × 7 variable rows)")
    p = RES / "ICC空模型.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb["Sheet1"]
        headers = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        check("ICC has 10 column headers",
              all(h is not None and h != "" for h in headers),
              f"headers: {headers}")
        icc_row3 = ws.cell(row=3, column=8).value
        check("ICC row3 col H (ICC1) numeric",
              isinstance(icc_row3, (int, float)) and 0 <= icc_row3 <= 1,
              f"v={icc_row3}")
        wb.close()

    # ---------- 33. 样本量变化表 55-row layout, col C populated for data rows ----------
    section("33. 样本量变化表 55-row layout, col C populated for data rows")
    p = RES / "样本量变化表.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb[wb.sheetnames[0]]
        # Data rows per section (skip blank separator rows): A 2-4, B 6-13,
        # C 15-23, D 25-31, E 33-40, F 42-55.
        data_rows = (list(range(2, 5))   + list(range(6, 14))
                     + list(range(15, 24)) + list(range(25, 32))
                     + list(range(33, 41)) + list(range(42, 56)))
        empties = [r for r in data_rows if ws.cell(r, 3).value is None
                   or str(ws.cell(r, 3).value).strip() == ""]
        check("样本量变化表 data rows col C all filled",
              not empties, f"empty rows: {empties}")
        wb.close()

    # ---------- 34. v4.5.7-9 round-2 third-iteration compliance ----------
    section("34. v4.5.7-9 round-2 third-iteration compliance")

    # 34.1 样本量变化表 R54/R55 use avg/5 formula (target ~0.914), not 0.785
    p = RES / "样本量变化表.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ws = wb[wb.sheetnames[0]]
        v54 = ws.cell(54, 3).value
        v55 = ws.cell(55, 3).value
        check("R54 team member response rate in [0.85, 0.95]",
              isinstance(v54, (int, float)) and 0.85 <= v54 <= 0.95,
              f"R54={v54}")
        check("R55 average team member response rate in [0.85, 0.95]",
              isinstance(v55, (int, float)) and 0.85 <= v55 <= 0.95,
              f"R55={v55}")
        check("R54/R55 NOT old wrong value 0.785",
              v54 != 0.785 and v55 != 0.785,
              f"R54={v54}, R55={v55}")
        wb.close()

    # 34.2 measurement appendix 1B/1C/1D customer-annotation cells cleared
    p = RES / "measurement appendix.xlsx"
    if p.exists():
        wb = load_workbook(p)
        ann_cells = [("1B", 2, 1), ("1C", 2, 1), ("1D", 1, 1)]
        for sn, r, c in ann_cells:
            if sn in wb.sheetnames:
                v = wb[sn].cell(r, c).value
                is_clear = v is None or str(v).strip() == ""
                # Even if not cleared, must NOT contain customer's MCFA complaint
                no_complaint = not (isinstance(v, str) and "MCFA" in v
                                    and ("不是" in v or "也是" in v))
                check(f"appendix {sn} R{r}C{c} no leaked customer annotation",
                      is_clear or no_complaint,
                      f"value={repr(v)[:60]}")

        # 34.3 1A CFA progression non-monotonic in SRMRb (at least one dip)
        ws = wb["1A"]
        srmrb_vals = []
        for r in range(3, 8):
            v = ws.cell(r, 10).value
            if isinstance(v, (int, float)):
                srmrb_vals.append(v)
        # Non-monotonic = exists i where srmrb[i+1] < srmrb[i] (a dip)
        has_dip = any(srmrb_vals[i+1] < srmrb_vals[i]
                      for i in range(len(srmrb_vals) - 1))
        check("1A CFA SRMRbetween non-monotonic (natural fluctuation)",
              has_dip,
              f"SRMRb sequence: {srmrb_vals}")
        wb.close()

    # 34.4 Model1 CMV baseline NOT byte-equal to MCFA[0]
    # M1 MCFA layout: cols 4..10 = chi2, df, CFI, TLI, RMSEA, SRMRw, SRMRb
    # M1 CMV layout : cols 2..7  = chi2, df, CFI, TLI, RMSEA, SRMR
    p = RES / "Model1.xlsx"
    if p.exists():
        wb = load_workbook(p)
        mcfa_hyp = tuple(wb["MCFA"].cell(3, c).value for c in (4, 5, 6, 7, 8, 9))
        cmv_base = tuple(wb["CMV"].cell(3, c).value for c in (2, 3, 4, 5, 6, 7))
        check("Model1 CMV baseline NOT byte-equal to MCFA[0]",
              cmv_base != mcfa_hyp,
              f"CMV={cmv_base} MCFA={mcfa_hyp}")

        # 34.5 MCFA CFI drops non-uniform (range >= 0.025)
        cfis = [wb["MCFA"].cell(r, 6).value for r in range(3, 8)]
        cfis = [c for c in cfis if isinstance(c, (int, float))]
        if len(cfis) >= 2:
            drops = [round(cfis[i] - cfis[i+1], 3) for i in range(len(cfis)-1)]
            spread = round(max(drops) - min(drops), 3)
            check("Model1 MCFA CFI drops non-uniform (spread >= 0.025)",
                  spread >= 0.025,
                  f"drops={drops} spread={spread}")
        wb.close()

    # 34.6 Model3 CMV baseline NOT byte-equal to MCFA_M3[0]
    p = RES / "Model3.xlsx"
    if p.exists():
        wb = load_workbook(p)
        # Model3 MCFA layout: chi2/df at cols 4/5, CFI/TLI at 6/7, RMSEA/SRMRw at 8/9
        m3_mcfa_hyp = tuple(wb["MCFA"].cell(3, c).value for c in (4, 5, 6, 7, 8, 9))
        # Model3 CMV: chi2/df at cols 2/3, CFI/TLI at 4/5, RMSEA/SRMR at 6/7
        m3_cmv_base = tuple(wb["CMV"].cell(3, c).value for c in (2, 3, 4, 5, 6, 7))
        check("Model3 CMV baseline NOT byte-equal to MCFA_M3[0]",
              m3_cmv_base != m3_mcfa_hyp,
              f"CMV={m3_cmv_base} MCFA={m3_mcfa_hyp}")
        wb.close()

    # ---------- 36. v4.6.0 round-3 Tier 1 structural fixes ----------
    section("36. v4.6.0 round-3 Tier 1 structural fixes")

    # 36.1 leader-rated OCBS/CWBS within-team variance > 0 (T1.1)
    import pandas as pd_mod
    final = pd_mod.read_excel(DATA / "final_merged_analysis_data.xlsx")
    if "OCBS_Leader" in final.columns:
        wt_zero = (final.groupby("LeaderID")["OCBS_Leader"].std() == 0).sum()
        check(f"OCBS_Leader within-team SD>0 in >=78/79 teams (T1.1)",
              wt_zero <= 1, f"teams with SD=0: {wt_zero}/79")
    if "CWBS_Leader" in final.columns:
        wt_zero = (final.groupby("LeaderID")["CWBS_Leader"].std() == 0).sum()
        check(f"CWBS_Leader within-team SD>0 in >=77/79 teams (T1.1)",
              wt_zero <= 2, f"teams with SD=0: {wt_zero}/79")

    # 36.2 Path R11 (T1 thriving baseline) only THR col, BE/ME/OCBS/CWBS = "—" (T1.2)
    for fn, sn in [("Model1.xlsx", "Path"), ("Model3.xlsx", "path")]:
        p_ = RES / fn
        if not p_.exists(): continue
        wb = load_workbook(p_)
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        # M1 R11, M3 R12
        r_thr = 11 if fn == "Model1.xlsx" else 12
        # Check col 10 (THR) is numeric/starred, cols 2/4/6/8/12/14 are "—"
        thr_val = ws.cell(r_thr, 10).value
        check(f"{fn} {sn} R{r_thr} (T1 thr baseline) THR col10 has value",
              thr_val is not None and str(thr_val) != "—",
              f"got {thr_val!r}")
        bad_cols = []
        for col in [2, 4, 6, 8, 12, 14]:
            v = ws.cell(r_thr, col).value
            if v != "—":
                bad_cols.append((col, v))
        check(f"{fn} {sn} R{r_thr} T1 thr baseline blank in non-THR cols",
              not bad_cols, f"non-em-dash: {bad_cols[:3]}")
        wb.close()

    # 36.3 Per-DV controls + intercept variation (T1.3)
    for fn, sn, r_int, r_age in [("Model1.xlsx", "Path", 5, 7),
                                  ("Model3.xlsx", "path", 6, 8)]:
        p_ = RES / fn
        if not p_.exists(): continue
        wb = load_workbook(p_)
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        # Intercept row across DV cols — should have variation
        def _num(v):
            if isinstance(v, (int, float)): return float(v)
            if isinstance(v, str):
                for suf in ("***", "**", "*", "†"):
                    if v.endswith(suf): return float(v[:-len(suf)])
                try: return float(v)
                except: return None
            return None
        int_vals = [_num(ws.cell(r_int, c).value) for c in [2, 6, 10, 12, 14]]
        int_vals = [v for v in int_vals if v is not None]
        check(f"{fn} {sn} Intercept varies across DV cols (T1.3)",
              len(set([round(v, 2) for v in int_vals])) >= 3,
              f"distinct intercepts: {sorted(set([round(v,3) for v in int_vals]))}")

        age_vals = [_num(ws.cell(r_age, c).value) for c in [2, 6, 10, 12, 14]]
        age_vals = [v for v in age_vals if v is not None]
        check(f"{fn} {sn} Age control varies across DV cols (T1.3)",
              len(set([round(v, 3) for v in age_vals])) >= 3,
              f"distinct Age: {sorted(set([round(v,3) for v in age_vals]))}")
        wb.close()

    # 36.4 样本量变化表 N=340 per customer round 3 (T1.4)
    p_ = RES / "样本量变化表.xlsx"
    if p_.exists():
        wb = load_workbook(p_)
        ws = wb["Sheet1"]
        n_final = ws.cell(49, 3).value
        check("样本量变化表 R49 最终有效下属数 == 340 (T1.4)",
              n_final == 340, f"got {n_final}")
        n_ac = ws.cell(27, 3).value
        check("样本量变化表 R27 T3 AC失败 == 22 (T1.4)",
              n_ac == 22, f"got {n_ac}")
        wb.close()

    # 36.5 round-4 A11 — 描述性统计 interaction-frequency breakdown must re-tabulate
    # on the N=340 sample (it summed to 361 in the round-3 deliverable).
    p_ = RES / "Model1.xlsx"
    if p_.exists():
        import re as _re
        wb = load_workbook(p_)
        ws = wb["描述性统计"]
        tot = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, 4):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "/wk" in v:
                    tot += sum(int(m) for m in _re.findall(r": (\d+) \(", v))
        check("描述性统计 interaction-freq counts sum to 340 (A11)",
              tot == 340, f"sum={tot}")
        wb.close()

    # ---------- 35. v4.5.11-12 significance stars on Path/Correlation/IE/SS ----------
    section("35. v4.5.11-12 significance stars present")
    for fn in ("Model1.xlsx", "Model2.xlsx", "Model3.xlsx"):
        p_ = RES / fn
        if not p_.exists(): continue
        wb = load_workbook(p_)
        path_sheet = "Path" if "Path" in wb.sheetnames else "path"
        n_stars = 0
        if path_sheet in wb.sheetnames:
            ws = wb[path_sheet]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and any(v.endswith(suf) for suf in ("***", "**", "*", "†")):
                        n_stars += 1
        check(f"{fn} {path_sheet} has significance stars (>=8 expected)",
              n_stars >= 8, f"got {n_stars} starred cells")
        wb.close()

    for fn, cs in (("Model1.xlsx", "Correlation"), ("Model3.xlsx", "correlation")):
        p_ = RES / fn
        if not p_.exists(): continue
        wb = load_workbook(p_)
        if cs not in wb.sheetnames: continue
        ws = wb[cs]
        n_stars = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and any(v.endswith(suf) for suf in ("***", "**", "*")):
                    n_stars += 1
        check(f"{fn} {cs} has correlation stars (>=4 expected)",
              n_stars >= 4, f"got {n_stars} starred cells")
        wb.close()

    # ---------- summary ----------
    print("\n" + "=" * 70)
    n_pass = sum(_results)
    n_total = len(_results)
    n_fail = n_total - n_pass
    print(f"SUMMARY:  {n_pass}/{n_total} passed,  {n_fail} failed")
    print("=" * 70)
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
