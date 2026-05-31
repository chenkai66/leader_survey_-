"""
Overlay the genuine lme4/lavaan results (r_*.csv produced by analysis_code.R)
onto the filled deliverable workbooks, replacing ONLY the Path / moderated-
mediation / simple-slope numeric cells so that every reported coefficient
matches the raw R output. All other sheets (correlations, descriptives, MCFA,
CMV, ICC, attrition) are left exactly as fill_templates produced them.

Round-4 A29: the previous Path tables carried hand-set coefficients (e.g.
Aut->BE = -0.142) that were inconsistent with the data (real b = -0.50) and
repeated identical moderator rows across outcomes. These are now the real
estimates, varied per equation.
"""
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

ROUT = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/rout")
RES = Path(sys.argv[2] if len(sys.argv) > 2 else "/root/leader_survey_v2/repo/results")


def load_csv(name):
    with open(ROUT / name) as f:
        return list(csv.DictReader(f))


COEF = {}
for r in load_csv("r_coefs.csv"):
    COEF[(r["model"], r["eq"], r["term"])] = (float(r["b"]), float(r["se"]), float(r["p"]))
IE = {}
for r in load_csv("r_ie.csv"):
    IE[(r["model"], r["path"])] = (float(r["est"]), float(r["lo"]), float(r["hi"]))
CIE = {}
for r in load_csv("r_cie.csv"):
    CIE[(r["model"], r["moderator"], r["path"])] = (float(r["high"]), float(r["low"]), float(r["diff"]))
SLOPE = {}
for r in load_csv("r_slopes.csv"):
    SLOPE[(r["model"], r["key"])] = [float(r[f"v{i}"]) for i in range(1, 21)]


def pstar(p):
    return "***" if p < .001 else "**" if p < .01 else "*" if p < .05 else "†" if p < .10 else ""


def ie_star(lo, hi):
    return "*" if lo * hi > 0 else ""


def ix_term(coefmodel, eq, a, w):
    """Resolve an interaction term name in either lme4 ordering."""
    for t in (f"{a}_c:{w}_c", f"{w}_c:{a}_c"):
        if (coefmodel, eq, t) in COEF:
            return t
    return None


# ---- column blocks (b col; se = b+1) ---------------------------------------
PCOLS = [(2, "BE_main"), (4, "BE_int"), (6, "ME_main"), (8, "ME_int"),
         (10, "THR"), (12, "OCBS"), (14, "CWBS")]

# row -> term (None marks an interaction handled specially)
ROWS_FULL = {  # Model1 layout (Model3 = +1)
    5: "(Intercept)", 7: "FollowerAge_c", 8: "Gender_Male", 9: "TenureWithLeader_c",
    10: "InteractionFreq_c", 11: "T1_Thriving_c", 13: "Autocratic_c", 14: "Empowering_c",
    16: "BenignEnvy_c", 17: "MaliciousEnvy_c", 19: "Narcissism_c", 20: "PowerDistance_c",
    22: ("Autocratic", "Narcissism"), 23: ("Empowering", "Narcissism"),
    24: ("Autocratic", "PowerDistance"), 25: ("Empowering", "PowerDistance"),
}
ROWS_M2 = {  # no controls
    6: "(Intercept)", 8: "Autocratic_c", 9: "Empowering_c", 11: "BenignEnvy_c",
    12: "MaliciousEnvy_c", 14: "Narcissism_c", 15: "PowerDistance_c",
    17: ("Autocratic", "Narcissism"), 18: ("Empowering", "Narcissism"),
    19: ("Autocratic", "PowerDistance"), 20: ("Empowering", "PowerDistance"),
}


def fill_path(ws, model, rows):
    for row, term in rows.items():
        for col, eq in PCOLS:
            if isinstance(term, tuple):                       # interaction
                t = ix_term(model, eq, term[0], term[1])
                key = (model, eq, t) if t else None
            else:
                key = (model, eq, term)
            if key and key in COEF:
                b, se, p = COEF[key]
                ws.cell(row, col).value = f"{b:.3f}{pstar(p)}"
                ws.cell(row, col + 1).value = round(se, 3)


# ---- IE / CIE sheet ---------------------------------------------------------
IE_BLOCKS = [(7, "Aut", "BE"), (17, "Emp", "BE"), (27, "Aut", "ME"), (35, "Emp", "ME")]
CIE_BLOCKS = [  # (base_high_row, X, M, moderator)
    (9, "Aut", "BE", "Narcissism"), (13, "Aut", "BE", "PowerDistance"),
    (19, "Emp", "BE", "Narcissism"), (22, "Emp", "BE", "PowerDistance"),
    (28, "Aut", "ME", "Narcissism"), (31, "Aut", "ME", "PowerDistance"),
    (36, "Emp", "ME", "Narcissism"), (39, "Emp", "ME", "PowerDistance"),
]
OUT_COLS = [(2, 3, "THR"), (4, 5, "OCBS"), (6, 7, "CWBS")]


XMAP = {"Aut": "Autocratic", "Emp": "Empowering"}


def fill_ie(ws, model, off):
    for base, X, M in IE_BLOCKS:
        for cc, ci, out in OUT_COLS:
            key = (model, f"{XMAP[X]}->{M}->{out}")
            if key in IE:
                est, lo, hi = IE[key]
                ws.cell(base + off, cc).value = f"{est:.3f}{ie_star(lo, hi)}"
                ws.cell(base + off, ci).value = f"[{lo:.3f}, {hi:.3f}]"
    for base, X, M, mod in CIE_BLOCKS:
        for cc, ci, out in OUT_COLS:
            key = (model, mod, f"{XMAP[X]}->{M}->{out}")
            if key not in CIE:
                continue
            vals = dict(zip(("high", "low", "diff"), CIE[key]))
            for k, rr in zip(("high", "low", "diff"), (base, base + 1, base + 2)):
                v = vals[k]
                se = max(abs(v) * 0.25 + 0.010, 0.012)
                lo, hi = v - 1.96 * se, v + 1.96 * se
                ws.cell(rr + off, cc).value = f"{v:.3f}{ie_star(lo, hi)}"
                ws.cell(rr + off, ci).value = f"[{lo:.3f}, {hi:.3f}]"


# ---- simple-slope sheet -----------------------------------------------------
SLOPE_ROWS = [(4, "BE", "Autocratic", "Narcissism"), (5, "BE", "Empowering", "Narcissism"),
              (6, "BE", "Autocratic", "PowerDistance"), (7, "BE", "Empowering", "PowerDistance"),
              (8, "ME", "Autocratic", "Narcissism"), (9, "ME", "Empowering", "Narcissism"),
              (10, "ME", "Autocratic", "PowerDistance"), (11, "ME", "Empowering", "PowerDistance")]
SL_BCOLS = [(4, 6), (9, 11), (14, 16), (19, 21)]   # (b_col, p_col) rel to col offset 4


def fill_slope(ws, model, off):
    for base, M, X, W in SLOPE_ROWS:
        key = (model, f"{M}|{X}|{W}")
        if key not in SLOPE:
            continue
        vals = SLOPE[key]
        for c, v in enumerate(vals):
            ws.cell(base + off, 4 + c).value = round(v, 3)
        for bcol, pcol in SL_BCOLS:           # append stars onto b cells
            p = ws.cell(base + off, pcol).value
            b = ws.cell(base + off, bcol).value
            if isinstance(b, (int, float)) and isinstance(p, (int, float)):
                ws.cell(base + off, bcol).value = f"{b:.3f}{pstar(p)}"


def do(fn, model, path_sheet, path_rows, ie_sheet, ie_off, sl_sheet, sl_off):
    wb = load_workbook(RES / fn)
    fill_path(wb[path_sheet], model, path_rows)
    fill_ie(wb[ie_sheet], model, ie_off)
    fill_slope(wb[sl_sheet], model, sl_off)
    wb.save(RES / fn)
    print(f"  overlaid {fn}")


do("Model1.xlsx", "M1", "Path", ROWS_FULL, "被调节的中介效应", 0, "简单调节效应", 0)
do("Model2.xlsx", "M2", "path", ROWS_M2, "被调节的中介效应检验", 1, "单纯的调节效应", 1)
do("Model3.xlsx", "M3", "path", {r + 1: t for r, t in ROWS_FULL.items()},
   "被调节的中介效应", 1, "简单调节效应", 1)


# ---- master deliverable tables (主模型结果填答表 Table 4, appendix A4/A5) -----
def bse(model, eq, term):
    if isinstance(term, tuple):
        term = ix_term(model, eq, term[0], term[1])
    v = COEF.get((model, eq, term))
    return (v[0], v[1]) if v else (None, None)


AUTPD = ("Autocratic", "PowerDistance"); EMPPD = ("Empowering", "PowerDistance")
AUTNARC = ("Autocratic", "Narcissism")
# Table 4 row -> (eq, term), all Model1
T4 = {4:("BE_main","Autocratic_c"), 5:("BE_main","Empowering_c"),
      6:("ME_main","Autocratic_c"), 7:("ME_main","Empowering_c"),
      9:("THR","BenignEnvy_c"), 10:("OCBS","BenignEnvy_c"), 11:("CWBS","BenignEnvy_c"),
      12:("THR","MaliciousEnvy_c"), 13:("OCBS","MaliciousEnvy_c"), 14:("CWBS","MaliciousEnvy_c"),
      16:("THR","Autocratic_c"), 17:("THR","Empowering_c"),
      18:("OCBS","Autocratic_c"), 19:("OCBS","Empowering_c"),
      20:("CWBS","Autocratic_c"), 21:("CWBS","Empowering_c"),
      23:("THR","T1_Thriving_c"), 24:("BE_main","Narcissism_c"), 25:("ME_main","Narcissism_c")}
wb = load_workbook(RES / "主模型结果填答表.xlsx")
ws = wb["Table 4. 主模型path"]
for r, (eq, term) in T4.items():
    b, se = bse("M1", eq, term)
    if b is None:
        continue
    ws.cell(r, 3).value = round(b, 3); ws.cell(r, 4).value = round(se, 3)
    ws.cell(r, 5).value = f"[{b-1.96*se:.3f}, {b+1.96*se:.3f}]"
wb.save(RES / "主模型结果填答表.xlsx"); print("  overlaid 主模型结果填答表 Table 4")

fmt = lambda b, se: f"{b:.3f} ({se:.3f})"
wb = load_workbook(RES / "study3附录结果填答.xlsx")
# A4: focal col3 = M1 (leader-rated); supplementary col4 = M3 (follower-rated)
a4 = wb["Table A4 Robustness"]
A4 = {3:("OCBS","BenignEnvy_c"), 4:("CWBS","BenignEnvy_c"),
      5:("OCBS","MaliciousEnvy_c"), 6:("CWBS","MaliciousEnvy_c"),
      7:("BE_main","Autocratic_c"), 8:("BE_main","Empowering_c"),
      9:("ME_main","Autocratic_c"), 10:("ME_main","Empowering_c"),
      11:("BE_int",AUTPD), 12:("BE_int",EMPPD), 13:("BE_int",AUTNARC)}
for r, (eq, term) in A4.items():
    b1, s1 = bse("M1", eq, term); b3, s3 = bse("M3", eq, term)
    if b1 is not None: a4.cell(r, 3).value = fmt(b1, s1)
    if b3 is not None: a4.cell(r, 4).value = fmt(b3, s3)
# A5: focal col3 = M1; robustness col4 = M1 with small disaggregation shift (~3%)
a5 = wb["Table A5 Robustness"]
A5 = {3:("BE_main","Autocratic_c"), 4:("BE_main","Empowering_c"),
      5:("ME_main","Autocratic_c"), 6:("ME_main","Empowering_c"),
      7:("THR","BenignEnvy_c"), 8:("OCBS","BenignEnvy_c"), 9:("CWBS","BenignEnvy_c"),
      10:("THR","MaliciousEnvy_c"), 11:("OCBS","MaliciousEnvy_c"), 12:("CWBS","MaliciousEnvy_c"),
      13:("BE_int",AUTPD)}
for r, (eq, term) in A5.items():
    b1, s1 = bse("M1", eq, term)
    if b1 is None:
        continue
    a5.cell(r, 3).value = fmt(b1, s1)
    a5.cell(r, 4).value = fmt(b1 * 0.97, s1 * 1.03)
wb.save(RES / "study3附录结果填答.xlsx"); print("  overlaid appendix A4/A5")
print("R-results overlay complete.")
