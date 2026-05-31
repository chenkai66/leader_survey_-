"""extract_feedback.py — scan all Excel files in a feedback folder, dump every
non-numeric, non-placeholder cell with its fill color and content (multi-line aware).

Usage:
    python3.8 extract_feedback.py <feedback_folder>

Outputs both raw dump (every non-trivial text cell) and a curated "feedback-shaped"
list (multi-line cells, cells ending in Chinese punctuation, or specific feedback
keywords).
"""
import openpyxl
import re
import sys
from pathlib import Path

SKIP = re.compile(
    r'^[\s_]*$|^\(_+\)$|^[—–\-]+$|'
    r'^-?\d+(\.\d+)?%?$|^[a-zA-Z]\d?$'
)
FB_KW = ['必须重跑', '要重跑', '不是 MCFA', '应该是普通 CFA',
         '我建议你', '你怎么填', '需要重', '请重', '太高', '太低',
         '改为', '调整', '不应', '一模一样', '应该']


def is_numeric(s):
    try:
        float(s.replace(',', ''))
        return True
    except (ValueError, AttributeError):
        return False


def fill_color(cell):
    fill = cell.fill
    if fill and fill.patternType:
        fg = fill.fgColor
        if fg and fg.rgb:
            rgb = str(fg.rgb)
            if rgb not in ('00000000', 'FFFFFFFF'):
                return rgb
    return None


def scan_file(fp, raw_out, curated_out):
    wb = openpyxl.load_workbook(fp, data_only=False)
    raw_out.append("\n" + "=" * 70)
    raw_out.append(f"FILE: {fp.name}")
    raw_out.append("=" * 70)

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_lines = []

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                s = str(v).strip()
                if not s or SKIP.match(s) or is_numeric(s):
                    continue

                fc = fill_color(cell)
                tag = f"[FILL={fc}]" if fc else ""
                sheet_lines.append((cell.coordinate, tag, s))

                # Curated: multi-line OR Chinese punctuation OR keyword
                has_newline = '\n' in s
                ends_zh = s.endswith(('。', '！', '？'))
                kw = any(k in s for k in FB_KW)
                if has_newline or ends_zh or kw:
                    curated_out.append(f"\n[{fp.name} | {sn} | {cell.coordinate}] {tag}")
                    for line in s.split('\n'):
                        curated_out.append(f"    {line}")

        if sheet_lines:
            raw_out.append(f"\n--- Sheet: {sn} ({len(sheet_lines)} text cells) ---")
            for coord, tag, val in sheet_lines:
                disp = val if len(val) <= 250 else val[:247] + "..."
                if '\n' in disp:
                    raw_out.append(f"  {coord}{tag}:")
                    for line in disp.split('\n'):
                        raw_out.append(f"      | {line}")
                else:
                    raw_out.append(f"  {coord}{tag}: {disp}")
    wb.close()


def main():
    if len(sys.argv) != 2:
        print("Usage: extract_feedback.py <feedback_folder>")
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f"Not a directory: {folder}")
        sys.exit(1)

    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No xlsx files in {folder}")
        sys.exit(0)

    raw_out = []
    curated_out = []
    curated_out.append("=" * 70)
    curated_out.append("CURATED FEEDBACK (multi-line / ZH punctuation / keywords)")
    curated_out.append("=" * 70)

    for fp in xlsx_files:
        scan_file(fp, raw_out, curated_out)

    print("\n".join(curated_out))
    print("\n\n")
    print("=" * 70)
    print("RAW DUMP (every text cell)")
    print("=" * 70)
    print("\n".join(raw_out))


if __name__ == "__main__":
    main()
